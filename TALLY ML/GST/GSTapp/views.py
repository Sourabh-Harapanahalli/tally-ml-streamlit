from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook,load_workbook
from io import BytesIO
import pandas as pd
import json
import pandas as pd
import numpy as np
from functools import reduce
import re
from django.views import View
from openpyxl import load_workbook
import io
import zipfile
from io import StringIO
from openpyxl.styles import Border,PatternFill,Side,Alignment
from datetime import datetime
from django.urls import reverse

# Create your views here.
def index(request):
    return render(request, 'GSTapp/index.html')

def tallyxml(request):
    return render(request,'GSTapp/tallyxml.html')

# def master_ledger(request):
#     return render(request,'GSTapp/master_ledger.html')

# def master_duties(request):
    # return render(request,'GSTapp/master_duties.html')

def master(request):
    return render(request,'GSTapp/master.html')

def jsonexcel(request):
    return render(request,'GSTapp/jsonexcel.html')

def download_excel(request):
    # Create a new Workbook

    template=pd.DataFrame(columns=('Supplier_Invoice','Datetime','Vch_Type','PartyLedgerName','Dr_LedgerName', 'Total_Amount', 
                                   'GST_5', 'GST_12','GST_18', 'GST_28', 'GST_5_IGST', 'GST_12_IGST',
                                   'GST_18_IGST','GST_28_IGST','Narration'))
    master_link=pd.DataFrame(data={'MASTER_LEDGER_COLUMN_LINK':['GST_5', 'GST_12','GST_18', 'GST_28',
                                      'GST_5_IGST', 'GST_12_IGST','GST_18_IGST','GST_28_IGST',
                                      '2.5_CGST', '6_CGST', '9_CGST','14_CGST', 
                                      '2.5_SGST', '6_SGST', '9_SGST', '14_SGST',
                                      '5_IGST','12_IGST', '18_IGST', '28_IGST'],'MASTER_LEDGER_TALLY_NAME':''})

    with pd.ExcelWriter('Template_Purchase.xlsx') as writer:
        template.to_excel(writer,index=False, sheet_name='TEMPLATE')
        master_link.to_excel(writer,index=False, sheet_name='MASTER_LEDGER_NAME_LINK')

    #Loading the Workbook
    wb_style = load_workbook('Template_Purchase.xlsx')

    #Accessing Product Informaiton Sheet
    sheet = wb_style.active

    #Specifying border style and color = red
    thick = Side(border_style="thin", color="000000")
    # for rows in sheet.iter_cols(min_col=16, max_col=None, min_row=1, max_row=100):
    #     for cell in rows:
    #         cell.fill = PatternFill(start_color="C1CDCD", end_color="C1CDCD",fill_type = "solid")
    #         cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    for sheet in wb_style.worksheets:
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 15
    wb_style.save("Template_Purchase.xlsx")

    # Create an HttpResponse with the workbook content as an attachment
    with open("Template_Purchase.xlsx", 'rb') as excel_file:
            # Create an HttpResponse with the file content as an attachment
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Template_Purchase.xlsx"'
        return response

def download_excel_pay_con_rec(request):
    # Create a new Workbook

    template=pd.DataFrame(columns=('DATE_TIME','PartyLedgerName','Dr_LedgerName','Dr_Amount','Cr_LedgerName','Cr_Amount','Narration','Vch_Type','NARRATION_1','NARRATION_2'))

    with pd.ExcelWriter('Template_Pay_Con_Rec.xlsx') as writer:
        template.to_excel(writer,index=False, sheet_name='TEMPLATE')

    #Loading the Workbook
    wb_style = load_workbook('Template_Pay_Con_Rec.xlsx')

    #Accessing Product Informaiton Sheet
    sheet = wb_style.active

    #Specifying border style and color = red
    thick = Side(border_style="thin", color="000000")
    # for rows in sheet.iter_cols(min_col=16, max_col=None, min_row=1, max_row=100):
    #     for cell in rows:
    #         cell.fill = PatternFill(start_color="C1CDCD", end_color="C1CDCD",fill_type = "solid")
    #         cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    for sheet in wb_style.worksheets:
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 15
    wb_style.save("Template_Pay_Con_Rec.xlsx")

    # Create an HttpResponse with the workbook content as an attachment
    with open("Template_Pay_Con_Rec.xlsx", 'rb') as excel_file:
            # Create an HttpResponse with the file content as an attachment
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Template_Pay_Con_Rec.xlsx"'
        return response


def download_excel_master_1(request):
    # Create a new Workbook

    template=pd.DataFrame(columns=('Ledger_Name','Alias','Group_Name','Country','State_Name','Pincode','Registration_Type','GST_NO','Opening_Balance','Dr/Cr','Address'))

    reference=pd.DataFrame(data={'Group_Name':['Branch / Divisions','Capital Account','Reserves & Surplus','Current Assets',
                                               'Bank Accounts','Cash-in-Hand','Deposits (Asset)','Loans & Advances (Asset)',
                                               'Stock-in-Hand','Sundry Debtors','Workers Advance','Current Liabilities','Provisions',
                                               'Sundry Creditors','S.CRS.EXP.PAYABLE','Direct Expenses','Direct Incomes',
                                               'Fixed Assets','Indirect Expenses','Indirect Incomes','Investments',
                                               'Loans (Liability)','Bank OD A/c','Secured Loans','Unsecured Loans',
                                               'Misc. Expenses (ASSET)','Suspense A/c']})

    with pd.ExcelWriter('Template_Master_Ledger.xlsx') as writer:
        template.to_excel(writer,index=False, sheet_name='TEMPLATE')
        reference.to_excel(writer,index=False, sheet_name='REFERENCE')

    #Loading the Workbook
    wb_style = load_workbook('Template_Master_Ledger.xlsx')

    # Create an HttpResponse with the workbook content as an attachment
    with open("Template_Master_Ledger.xlsx", 'rb') as excel_file:
            # Create an HttpResponse with the file content as an attachment
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Template_Master_Ledger.xlsx"'
        return response

def download_excel_master_2(request):
    # Create a new Workbook

    template=pd.DataFrame(columns=('Ledger_Name','Group_Name','Rate_of_Tax','Tax_Type'))
    reference=pd.DataFrame(data={'Group_Name':['Duties & Taxes','',''],'Tax_Type':['State Tax','Central Tax','Integrated Tax']})

    with pd.ExcelWriter('Template_Master_DPS.xlsx') as writer:
        template.to_excel(writer,index=False, sheet_name='TEMPLATE')
        reference.to_excel(writer,index=False, sheet_name='REFERENCE')

    #Loading the Workbook
    wb_style = load_workbook('Template_Master_DPS.xlsx')

    # Create an HttpResponse with the workbook content as an attachment
    with open("Template_Master_DPS.xlsx", 'rb') as excel_file:
            # Create an HttpResponse with the file content as an attachment
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Template_Master_DPS.xlsx"'
        return response
    
def download_excel_master_3(request):
    # Create a new Workbook

    template=pd.DataFrame(columns=('Ledger_Name','Group_Name','Nature_of_transaction','RATE_OF_CGST_SGST','RATE_OF_IGST'))
    reference=pd.DataFrame(data={'Group_Name':['Purchase Accounts','Sales Accounts','','','','','',''],
                                 'Nature_of_transaction':['Interstate Sales Exempt','Interstate Sales Taxable','Sales Exempt',
                                                          'Sales Taxable','Interstate Purchase Exempt','Interstate Purchase Taxable',
                                                          'Purchase Exempt','Purchase Taxable']})

    with pd.ExcelWriter('Template_Master_PS.xlsx') as writer:
        template.to_excel(writer,index=False, sheet_name='TEMPLATE')
        reference.to_excel(writer,index=False, sheet_name='REFERENCE')

    #Loading the Workbook
    wb_style = load_workbook('Template_Master_PS.xlsx')

    # Create an HttpResponse with the workbook content as an attachment
    with open("Template_Master_PS.xlsx", 'rb') as excel_file:
            # Create an HttpResponse with the file content as an attachment
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Template_Master_PS.xlsx"'
        return response


class Purchase_Sales(View):
    def get(self,request):
        return render(request,'GSTapp/pur_sale.html')
    
    def post(self,request):
        uploaded_file = request.FILES.get('upload_file')

        if uploaded_file:
            pd.set_option('future.no_silent_downcasting', True)
            data=pd.read_excel(uploaded_file,sheet_name='TEMPLATE')
            master=pd.read_excel(uploaded_file,sheet_name='MASTER_LEDGER_NAME_LINK',index_col='MASTER_LEDGER_COLUMN_LINK')
            
            # for x in data:
            #     if ('int' in str(data[x].dtypes) or 'float' in str(data[x].dtypes)) and 'Supplier_Invoice' not in x:
            #         data[x]=data[x].astype("float").round(2)
            numeric_cols = data.select_dtypes(include=['int64', 'float64']).columns
            filtered_cols = [col for col in numeric_cols if 'Supplier_Invoice' not in col]

            # Convert selected columns to float and round them to two decimal places
            data[filtered_cols] = data[filtered_cols].astype(float).round(2)
            
                
            data=data.fillna('')
            data['PartyLedgerName'] = data['PartyLedgerName'].str.replace('&', '&amp;', regex=False)
            data['Dr_LedgerName'] = data['Dr_LedgerName'].str.replace('&', '&amp;', regex=False)

            # data['GST_5'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_12'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_18'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_28'] = pd.to_numeric(data['GST_5'], errors='coerce')

            # data['GST_5_IGST'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_12_IGST'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_18_IGST'] = pd.to_numeric(data['GST_5'], errors='coerce')
            # data['GST_28_IGST'] = pd.to_numeric(data['GST_5'], errors='coerce')
            data['GST_5'] = data['GST_5'].replace('', np.nan)
            data['GST_12'] = data['GST_12'].replace('', np.nan)
            data['GST_18'] = data['GST_18'].replace('', np.nan)
            data['GST_28'] = data['GST_28'].replace('', np.nan)

            data['GST_5_IGST'] = data['GST_5_IGST'].replace('', np.nan)
            data['GST_12_IGST'] = data['GST_12_IGST'].replace('', np.nan)
            data['GST_18_IGST'] = data['GST_18_IGST'].replace('', np.nan)
            data['GST_28_IGST'] = data['GST_28_IGST'].replace('', np.nan)

            data['2.5_CGST']=data['GST_5'].astype(float)*(2.5/100)
            data['6_CGST']=data['GST_12'].astype(float)*(6/100)
            data['9_CGST']=data['GST_18'].astype(float)*(9/100)
            data['14_CGST']=data['GST_28'].astype(float)*(14/100)

            data['5_IGST']=data['GST_5_IGST'].astype(float)*(5/100)
            data['12_IGST']=data['GST_12_IGST'].astype(float)*(12/100)
            data['18_IGST']=data['GST_18_IGST'].astype(float)*(18/100)
            data['28_IGST']=data['GST_28_IGST'].astype(float)*(28/100)

            data['2.5_SGST']=data['2.5_CGST']
            data['6_SGST']=data['6_CGST']
            data['9_SGST']=data['9_CGST']
            data['14_SGST']=data['14_CGST']

            data['Dr_Amount']=data['Total_Amount'].round(0)

            # for x in data:
            #     if ('int' in str(data[x].dtypes) or 'float' in str(data[x].dtypes)) and 'Supplier_Invoice' not in x:
            #         data[x]=data[x].astype("float").round(2)
            numeric_cols_1 = data.select_dtypes(include=['int64', 'float64']).columns
            filtered_cols_1 = [col for col in numeric_cols_1 if 'Supplier_Invoice' not in col]

            data[filtered_cols_1] = data[filtered_cols_1].astype(float).round(2)

            data['CR_TOTAL']=data[['GST_5','GST_12',
            'GST_18',
            'GST_28',
            'GST_5_IGST',
            'GST_12_IGST',
            'GST_18_IGST',
            'GST_28_IGST',
            '2.5_CGST',
            '6_CGST',
            '9_CGST',
            '14_CGST',
            '2.5_SGST',
            '6_SGST',
            '9_SGST',
            '14_SGST',
            '5_IGST',
            '12_IGST',
            '18_IGST',
            '28_IGST',
            ]].sum(axis=1).round(2)


            def replace_negatives(x):
                if x < 0:
                    return 0
                else:
                    return x

            def replace_null(x):
                if x == 0:
                    return ''
                else:
                    return x

            data['Cr_AmountThree']=data['Dr_Amount']-data['CR_TOTAL']
            data['Cr_AmountThree']=data['Cr_AmountThree'].round(2).map(replace_negatives)
            data['Cr_AmountThree']=data['Cr_AmountThree'].round(2).map(replace_null)

            data['Dr_AmountOne']=data['CR_TOTAL']-data['Dr_Amount']
            data['Dr_AmountOne']=data['Dr_AmountOne'].round(2).map(replace_negatives)
            data['Dr_AmountOne']=data['Dr_AmountOne'].round(2).map(replace_null)
            

            data['Dr_LedgerNameOne'] = None
            data['Cr_LedgerNameThree'] = None
            # data['Diff'] = None
            
            data['Dr_LedgerNameOne'] = data['Dr_AmountOne'].apply(lambda x: 'ROUND OFF' if x != '' else '')

            # for x in range(len(data['Dr_AmountOne'])):
            #     if data['Dr_AmountOne'].iloc[x]!='':
            #         data['Dr_LedgerNameOne'].iloc[x]='ROUND OFF'
            #     elif data['Dr_AmountOne'].iloc[x]=='':
            #         data['Dr_LedgerNameOne'].iloc[x]=''

            data['Cr_LedgerNameThree'] = data['Cr_AmountThree'].apply(lambda x: 'ROUND OFF' if x != '' else '')


            # for x in range(len(data['Cr_AmountThree'])):
            #     if data['Cr_AmountThree'].iloc[x]!='':
            #         data['Cr_LedgerNameThree'].iloc[x]='ROUND OFF'
            #     elif data['Cr_AmountThree'].iloc[x]=='':
            #         data['Cr_LedgerNameThree'].iloc[x]=''

            data['Dr_(+/-)']=None
            data['ledger_info']=None
            data['ledger_isparty']=None
            data['master_info']=None
            data['master_isparty']=None

            data.loc[data['Vch_Type'] == 'Purchase', 'Dr_(+/-)'] = '-'
            data.loc[data['Vch_Type'] == 'Purchase', 'ledger_info'] = 'No'
            data.loc[data['Vch_Type'] == 'Purchase', 'ledger_isparty'] = 'Yes'
            data.loc[data['Vch_Type'] == 'Purchase', 'master_info'] = 'Yes'
            data.loc[data['Vch_Type'] == 'Purchase', 'master_isparty'] = 'No'
            
            data.loc[data['Vch_Type'] == 'Sales', 'Dr_(+/-)'] = '+'
            data.loc[data['Vch_Type'] == 'Sales', 'ledger_info'] = 'Yes'
            data.loc[data['Vch_Type'] == 'Sales', 'ledger_isparty'] = 'No'
            data.loc[data['Vch_Type'] == 'Sales', 'master_info'] = 'No'
            data.loc[data['Vch_Type'] == 'Sales', 'master_isparty'] = 'Yes'

            #print('datetime' in str(data.dtypes['Datetime']))
            # data['DATE'] = pd.NaT
            # if 'datetime' in str(data.dtypes['Datetime']):
            #     data['DATE']=data['DATE'].dt.strftime('%Y%m%d')
            # else:        
            #     for x in range(len(data['Datetime'])):
            #         if len(data['Datetime'].iloc[x].split('/'))==3:
            #             string=data['Datetime'][x].split('/')
            #             data['DATE'].iloc[x]=string[-1]+string[-2]+string[-3]
            #         elif len(data['Datetime'].iloc[x].split('/'))==3:
            #             string=data['Datetime'][x].split('-')
            #             data['DATE'].iloc[x]=string[-1]+string[-2]+string[-3]

            # try:
            #     data['DATE']=data['DATE'].dt.strftime('%Y%m%d')
            # except:
            #     data['DATE']

            # default_date = pd.to_datetime('1998-01-01')  # Define your default date
            # data['DATE'] = data['DATE'].fillna(default_date)

            try:
                data['DATE'] = pd.NaT
                # If 'Datetime' column is already datetime type, format it
                data['DATE'] = pd.to_datetime(data['Datetime'],dayfirst=True).dt.strftime('%Y%m%d')
            except:
                data['DATE'] = None
                # If 'Datetime' column is string type, parse and format it
                for x in range(len(data['Datetime'])):
                    if '/' in str(data['Datetime'].iloc[x]):
                        string = str(data['Datetime'].iloc[x]).split('/')
                    elif '-' in str(data['Datetime'].iloc[x]):
                        string = str(data['Datetime'].iloc[x]).split('-')
                    else:
                        continue
                    data['DATE'].iloc[x] = string[-1] + string[-2] + string[-3]

            data['DATE'] = data['DATE'].fillna('19980101')
            data=data.fillna('')
            data.to_excel("data.xlsx")

            # data['Diff']=(data['Dr_Amount']+data['Dr_AmountOne'].replace('', 0)-data['CR_TOTAL']-data['Cr_AmountThree'].replace('', 0)).round(4)

            xml_begin= '''
                <ENVELOPE>
                <HEADER>
                <TALLYREQUEST>Import Data</TALLYREQUEST>
                </HEADER>
                <BODY>
                <IMPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>All Masters</REPORTNAME>
                    <STATICVARIABLES>
                    <SVCURRENTCOMPANY>Demo Company</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
                <REQUESTDATA>
                '''
            xml_end='''
                </REQUESTDATA>
                </IMPORTDATA>
                </BODY>
                </ENVELOPE>
                '''
            xml=''

            for row in range(len(data)):


                xml=xml+'''<TALLYMESSAGE xmlns:UDF="TallyUDF">
                    <VOUCHER REMOTEID="" VCHKEY="" VCHTYPE="'''+str(data.iloc[row]['Vch_Type']).strip()+'''" ACTION="Create" OBJVIEW="Accounting Voucher View">
                    <DATE>'''+str(data.iloc[row]['DATE']).strip()+'''</DATE>
                    <REFERENCEDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</REFERENCEDATE>                                                                                              
                    <GUID></GUID>
                    <NARRATION>'''+str(data.iloc[row]['Narration']).strip()+'''</NARRATION>
                    <VOUCHERTYPENAME>'''+str(data.iloc[row]['Vch_Type']).strip()+'''</VOUCHERTYPENAME>
                    <REFERENCE>'''+str(data.iloc[row]['Supplier_Invoice']).strip()+'''</REFERENCE>
                    <VOUCHERNUMBER>1</VOUCHERNUMBER>
                    <PARTYLEDGERNAME>'''+str(data.iloc[row]['PartyLedgerName']).strip()+'''</PARTYLEDGERNAME>
                    <CSTFORMISSUETYPE/>
                    <CSTFORMRECVTYPE/>
                    <FBTPAYMENTTYPE>Default</FBTPAYMENTTYPE>
                    <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>
                    <VCHGSTCLASS/>
                    <DIFFACTUALQTY>No</DIFFACTUALQTY>
                    <ISMSTFROMSYNC>No</ISMSTFROMSYNC>
                    <ASORIGINAL>No</ASORIGINAL>
                    <AUDITED>No</AUDITED>
                    <FORJOBCOSTING>No</FORJOBCOSTING>
                    <ISOPTIONAL>No</ISOPTIONAL>
                    <EFFECTIVEDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</EFFECTIVEDATE>
                    <USEFOREXCISE>No</USEFOREXCISE>
                    <ISFORJOBWORKIN>No</ISFORJOBWORKIN>
                    <ALLOWCONSUMPTION>No</ALLOWCONSUMPTION>
                    <USEFORINTEREST>No</USEFORINTEREST>
                    <USEFORGAINLOSS>No</USEFORGAINLOSS>
                    <USEFORGODOWNTRANSFER>No</USEFORGODOWNTRANSFER>
                    <USEFORCOMPOUND>No</USEFORCOMPOUND>
                    <USEFORSERVICETAX>No</USEFORSERVICETAX>
                    <ISEXCISEVOUCHER>No</ISEXCISEVOUCHER>
                    <EXCISETAXOVERRIDE>No</EXCISETAXOVERRIDE>
                    <USEFORTAXUNITTRANSFER>No</USEFORTAXUNITTRANSFER>
                    <EXCISEOPENING>No</EXCISEOPENING>
                    <USEFORFINALPRODUCTION>No</USEFORFINALPRODUCTION>
                    <ISTDSOVERRIDDEN>No</ISTDSOVERRIDDEN>
                    <ISTCSOVERRIDDEN>No</ISTCSOVERRIDDEN>
                    <ISTDSTCSCASHVCH>No</ISTDSTCSCASHVCH>
                    <INCLUDEADVPYMTVCH>No</INCLUDEADVPYMTVCH>
                    <ISSUBWORKSCONTRACT>No</ISSUBWORKSCONTRACT>
                    <ISVATOVERRIDDEN>No</ISVATOVERRIDDEN>
                    <IGNOREORIGVCHDATE>No</IGNOREORIGVCHDATE>
                    <ISSERVICETAXOVERRIDDEN>No</ISSERVICETAXOVERRIDDEN>
                    <ISISDVOUCHER>No</ISISDVOUCHER>
                    <ISEXCISEOVERRIDDEN>No</ISEXCISEOVERRIDDEN>
                    <ISEXCISESUPPLYVCH>No</ISEXCISESUPPLYVCH>
                    <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>
                    <GSTNOTEXPORTED>No</GSTNOTEXPORTED>
                    <ISVATPRINCIPALACCOUNT>No</ISVATPRINCIPALACCOUNT>
                    <ISBOENOTAPPLICABLE>No</ISBOENOTAPPLICABLE>
                    <ISSHIPPINGWITHINSTATE>No</ISSHIPPINGWITHINSTATE>
                    <ISOVERSEASTOURISTTRANS>No</ISOVERSEASTOURISTTRANS>
                    <ISCANCELLED>No</ISCANCELLED>
                    <HASCASHFLOW>No</HASCASHFLOW>
                    <ISPOSTDATED>No</ISPOSTDATED>
                    <USETRACKINGNUMBER>No</USETRACKINGNUMBER>
                    <ISINVOICE>No</ISINVOICE>
                    <MFGJOURNAL>No</MFGJOURNAL>
                    <HASDISCOUNTS>No</HASDISCOUNTS>
                    <ASPAYSLIP>No</ASPAYSLIP>
                    <ISCOSTCENTRE>No</ISCOSTCENTRE>
                    <ISSTXNONREALIZEDVCH>No</ISSTXNONREALIZEDVCH>
                    <ISEXCISEMANUFACTURERON>No</ISEXCISEMANUFACTURERON>
                    <ISBLANKCHEQUE>No</ISBLANKCHEQUE>
                    <ISVOID>No</ISVOID>
                    <ISONHOLD>No</ISONHOLD>
                    <ORDERLINESTATUS>No</ORDERLINESTATUS>
                    <VATISAGNSTCANCSALES>No</VATISAGNSTCANCSALES>
                    <VATISPURCEXEMPTED>No</VATISPURCEXEMPTED>
                    <ISVATRESTAXINVOICE>No</ISVATRESTAXINVOICE>
                    <VATISASSESABLECALCVCH>No</VATISASSESABLECALCVCH>
                    <ISVATDUTYPAID>Yes</ISVATDUTYPAID>
                    <ISDELIVERYSAMEASCONSIGNEE>No</ISDELIVERYSAMEASCONSIGNEE>
                    <ISDISPATCHSAMEASCONSIGNOR>No</ISDISPATCHSAMEASCONSIGNOR>
                    <ISDELETED>No</ISDELETED>
                    <CHANGEVCHMODE>No</CHANGEVCHMODE>
                    <ALTERID> 1</ALTERID>
                    <MASTERID> 1</MASTERID>
                    <VOUCHERKEY>184322816475144</VOUCHERKEY>
                    <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                    <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                    <DUTYHEADDETAILS.LIST>      </DUTYHEADDETAILS.LIST>
                    <SUPPLEMENTARYDUTYHEADDETAILS.LIST>      </SUPPLEMENTARYDUTYHEADDETAILS.LIST>
                    <INVOICEDELNOTES.LIST>      </INVOICEDELNOTES.LIST>
                    <INVOICEORDERLIST.LIST>      </INVOICEORDERLIST.LIST>
                    <INVOICEINDENTLIST.LIST>      </INVOICEINDENTLIST.LIST>
                    <ATTENDANCEENTRIES.LIST>      </ATTENDANCEENTRIES.LIST>
                    <ORIGINVOICEDETAILS.LIST>      </ORIGINVOICEDETAILS.LIST>
                    <INVOICEEXPORTLIST.LIST>      </INVOICEEXPORTLIST.LIST>

                    <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(data.iloc[row]['Dr_LedgerName']).strip()+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['ledger_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['ledger_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['ledger_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+"{0:+}".format(int(data.iloc[0]['Dr_(+/-)']+"1")*-1)[0]+"{:.2f}".format(data.iloc[row]['Dr_Amount']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+"{0:+}".format(int(data.iloc[0]['Dr_(+/-)']+"1")*-1)[0]+"{:.2f}".format(data.iloc[row]['Dr_Amount']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>




                    <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_18'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_18']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_18']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_5'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_5']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_5']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>








                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_12'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_12']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_12']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>


                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_28'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_28']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_28']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>


                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['2.5_CGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['2.5_CGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['2.5_CGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['6_CGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['6_CGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['6_CGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['9_CGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['9_CGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['9_CGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['14_CGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['14_CGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['14_CGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['2.5_SGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['2.5_SGST']).strip()+''' </AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['2.5_SGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['6_SGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['6_SGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['6_SGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['9_SGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['9_SGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['9_SGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['14_SGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['14_SGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['14_SGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_5_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_5_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_5_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>







                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_12_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_12_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_12_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_18_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_18_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_18_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['GST_28_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_28_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['GST_28_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>


                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['5_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['5_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['5_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['12_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['12_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['12_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['18_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['18_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['18_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>

                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(master.loc['28_IGST'].iloc[0])+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['28_IGST']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['28_IGST']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>


                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="decimal">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(data.iloc[row]['Dr_LedgerNameOne']).strip()+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['ledger_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['ledger_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['ledger_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+"{0:+}".format(int(data.iloc[0]['Dr_(+/-)']+"1")*-1)[0]+str(data.iloc[row]['Dr_AmountOne']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+"{0:+}".format(int(data.iloc[0]['Dr_(+/-)']+"1")*-1)[0]+str(data.iloc[row]['Dr_AmountOne']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>





                <ALLLEDGERENTRIES.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <LEDGERNAME>'''+str(data.iloc[row]['Cr_LedgerNameThree']).strip()+'''</LEDGERNAME>
                    <GSTCLASS/>
                    <ISDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISDEEMEDPOSITIVE>
                    <LEDGERFROMITEM>No</LEDGERFROMITEM>
                    <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                    <ISPARTYLEDGER>'''+str(data.iloc[row]['master_isparty'])+'''</ISPARTYLEDGER>
                    <ISLASTDEEMEDPOSITIVE>'''+str(data.iloc[row]['master_info'])+'''</ISLASTDEEMEDPOSITIVE>
                    <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                    <AMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['Cr_AmountThree']).strip()+'''</AMOUNT>
                    <VATEXPAMOUNT>'''+str(data.iloc[row]['Dr_(+/-)'])+str(data.iloc[row]['Cr_AmountThree']).strip()+'''</VATEXPAMOUNT>
                    <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                    <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                    <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                    <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                    <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                    <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                    <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                    <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                    <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                    <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                    <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                    <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                    <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                    <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                    <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                    <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                    <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                    <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                    <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                    <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                    </ALLLEDGERENTRIES.LIST>



                    <PAYROLLMODEOFPAYMENT.LIST>      </PAYROLLMODEOFPAYMENT.LIST>
                    <ATTDRECORDS.LIST>      </ATTDRECORDS.LIST>
                    <TEMPGSTRATEDETAILS.LIST>      </TEMPGSTRATEDETAILS.LIST>
                    </VOUCHER>
                    </TALLYMESSAGE> '''

            xml_final=xml_begin+xml+xml_end
            request.session['xml_final'] = xml_final
            csv_string = data.to_csv(index=False)
            request.session['data_csv'] = csv_string
            request.session['file_name'] = 'Purchase_Sales'
        # You can add further processing of the uploaded file here

        return render(request, 'GSTapp/pur_sale.html', {'message':True,'file_name':uploaded_file})

class Master_Ledger(View):
    def get(self,request):
        return render(request,'GSTapp/master_ledger.html')
    
    def post(self,request):
        uploaded_file = request.FILES.get('upload_file_master_ledger')

        if uploaded_file:
            data=pd.read_excel(uploaded_file,sheet_name='TEMPLATE')
   
            data=data.fillna('')

            data['Dr_(+/-)']=None
            data.loc[data['Dr/Cr'] == 'Dr', 'Dr_(+/-)'] = -1
            data.loc[data['Dr/Cr'] == 'Cr', 'Dr_(+/-)'] = 1

            data['DATE']= datetime.today().strftime('%Y%m%d')

            data['OB']=data['Opening_Balance']*data['Dr_(+/-)']

            data.to_excel("output.xlsx",sheet_name='TEMPLATE') 

            def replace_amp(text):
                return text.replace('&', '&amp;')
            
            data['Ledger_Name'] = data['Ledger_Name'].apply(replace_amp)
            data['Address'] = data['Address'].apply(replace_amp) 
            data['State_Name'] = data['State_Name'].apply(replace_amp) 
            data['Country'] = data['Country'].apply(replace_amp) 
            data['Group_Name'] = data['Group_Name'].apply(replace_amp) 
            data['Alias'] = data['Alias'].apply(replace_amp)            

            xml_begin= '''
                <ENVELOPE>
                <HEADER>
                <TALLYREQUEST>Import Data</TALLYREQUEST>
                </HEADER>
                <BODY>
                <IMPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>All Masters</REPORTNAME>
                    <STATICVARIABLES>
                    <SVCURRENTCOMPANY>Demo Company</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
                <REQUESTDATA>
                '''
            xml_end='''
                </REQUESTDATA>
                </IMPORTDATA>
                </BODY>
                </ENVELOPE>
                '''
            xml=''

            for row in range(len(data)):
                xml=xml+'''
                <TALLYMESSAGE xmlns:UDF="TallyUDF">
                    <LEDGER NAME="'''+str(data.iloc[row]['Ledger_Name']).strip()+'''" RESERVEDNAME="">
                    <ADDRESS.LIST TYPE="String">
                    <ADDRESS>'''+str(data.iloc[row]['Address']).strip()+'''</ADDRESS>
                    </ADDRESS.LIST>
                    <MAILINGNAME.LIST TYPE="String">
                    <MAILINGNAME>'''+str(data.iloc[row]['Ledger_Name']).strip()+'''</MAILINGNAME>
                    </MAILINGNAME.LIST>
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                    <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <STARTINGFROM>'''+str(data.iloc[row]['DATE']).strip()+'''</STARTINGFROM>
                    <CREATEDDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</CREATEDDATE>
                    <ALTEREDON>'''+str(data.iloc[row]['DATE']).strip()+'''</ALTEREDON>
                    <GUID>5146d6da-e8bc-454f-8fd1-116f5719b3cc-00000248</GUID>
                    <CURRENCYNAME>₹</CURRENCYNAME>
                    <PRIORSTATENAME>'''+str(data.iloc[row]['State_Name']).strip()+'''</PRIORSTATENAME>
                    <PINCODE>'''+(str(int(data.iloc[row]['Pincode'])) if isinstance(data.iloc[row]['Pincode'], (int, float)) else '').strip()+'''</PINCODE>
                    <COUNTRYNAME>'''+str(data.iloc[row]['Country']).strip()+'''</COUNTRYNAME>
                    <GSTREGISTRATIONTYPE>Regular</GSTREGISTRATIONTYPE>
                    <VATDEALERTYPE>Regular</VATDEALERTYPE>
                    <PARENT>'''+str(data.iloc[row]['Group_Name']).strip()+'''</PARENT>
                    <CREATEDBY>amit</CREATEDBY>
                    <ALTEREDBY>amit</ALTEREDBY>
                    <TAXCLASSIFICATIONNAME/>
                    <TAXTYPE>Others</TAXTYPE>
                    <COUNTRYOFRESIDENCE>'''+str(data.iloc[row]['Country']).strip()+'''</COUNTRYOFRESIDENCE>
                    <GSTTYPE/>
                    <APPROPRIATEFOR/>
                    <PARTYGSTIN>'''+str(data.iloc[row]['GST_NO']).strip()+'''</PARTYGSTIN>
                    <LEDSTATENAME>'''+str(data.iloc[row]['State_Name']).strip()+'''</LEDSTATENAME>
                    <SERVICECATEGORY>&#4; Not Applicable</SERVICECATEGORY>
                    <EXCISELEDGERCLASSIFICATION/>
                    <EXCISEDUTYTYPE/>
                    <EXCISENATUREOFPURCHASE/>
                    <LEDGERFBTCATEGORY/>
                    <ISBILLWISEON>No</ISBILLWISEON>
                    <ISCOSTCENTRESON>No</ISCOSTCENTRESON>
                    <ISINTERESTON>No</ISINTERESTON>
                    <ALLOWINMOBILE>No</ALLOWINMOBILE>
                    <ISCOSTTRACKINGON>No</ISCOSTTRACKINGON>
                    <ISBENEFICIARYCODEON>No</ISBENEFICIARYCODEON>
                    <PLASINCOMEEXPENSE>No</PLASINCOMEEXPENSE>
                    <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
                    <ASORIGINAL>Yes</ASORIGINAL>
                    <ISCONDENSED>No</ISCONDENSED>
                    <AFFECTSSTOCK>No</AFFECTSSTOCK>
                    <ISRATEINCLUSIVEVAT>No</ISRATEINCLUSIVEVAT>
                    <FORPAYROLL>No</FORPAYROLL>
                    <ISABCENABLED>No</ISABCENABLED>
                    <ISCREDITDAYSCHKON>No</ISCREDITDAYSCHKON>
                    <INTERESTONBILLWISE>No</INTERESTONBILLWISE>
                    <OVERRIDEINTEREST>No</OVERRIDEINTEREST>
                    <OVERRIDEADVINTEREST>No</OVERRIDEADVINTEREST>
                    <USEFORVAT>No</USEFORVAT>
                    <IGNORETDSEXEMPT>No</IGNORETDSEXEMPT>
                    <ISTCSAPPLICABLE>No</ISTCSAPPLICABLE>
                    <ISTDSAPPLICABLE>No</ISTDSAPPLICABLE>
                    <ISFBTAPPLICABLE>No</ISFBTAPPLICABLE>
                    <ISGSTAPPLICABLE>No</ISGSTAPPLICABLE>
                    <ISEXCISEAPPLICABLE>No</ISEXCISEAPPLICABLE>
                    <ISTDSEXPENSE>No</ISTDSEXPENSE>
                    <ISEDLIAPPLICABLE>No</ISEDLIAPPLICABLE>
                    <ISRELATEDPARTY>No</ISRELATEDPARTY>
                    <USEFORESIELIGIBILITY>No</USEFORESIELIGIBILITY>
                    <ISINTERESTINCLLASTDAY>No</ISINTERESTINCLLASTDAY>
                    <APPROPRIATETAXVALUE>No</APPROPRIATETAXVALUE>
                    <ISBEHAVEASDUTY>No</ISBEHAVEASDUTY>
                    <INTERESTINCLDAYOFADDITION>No</INTERESTINCLDAYOFADDITION>
                    <INTERESTINCLDAYOFDEDUCTION>No</INTERESTINCLDAYOFDEDUCTION>
                    <ISOTHTERRITORYASSESSEE>No</ISOTHTERRITORYASSESSEE>
                    <OVERRIDECREDITLIMIT>No</OVERRIDECREDITLIMIT>
                    <ISAGAINSTFORMC>No</ISAGAINSTFORMC>
                    <ISCHEQUEPRINTINGENABLED>Yes</ISCHEQUEPRINTINGENABLED>
                    <ISPAYUPLOAD>No</ISPAYUPLOAD>
                    <ISPAYBATCHONLYSAL>No</ISPAYBATCHONLYSAL>
                    <ISBNFCODESUPPORTED>No</ISBNFCODESUPPORTED>
                    <ALLOWEXPORTWITHERRORS>No</ALLOWEXPORTWITHERRORS>
                    <CONSIDERPURCHASEFOREXPORT>No</CONSIDERPURCHASEFOREXPORT>
                    <ISTRANSPORTER>No</ISTRANSPORTER>
                    <USEFORNOTIONALITC>No</USEFORNOTIONALITC>
                    <ISECOMMOPERATOR>No</ISECOMMOPERATOR>
                    <SHOWINPAYSLIP>No</SHOWINPAYSLIP>
                    <USEFORGRATUITY>No</USEFORGRATUITY>
                    <ISTDSPROJECTED>No</ISTDSPROJECTED>
                    <FORSERVICETAX>No</FORSERVICETAX>
                    <ISINPUTCREDIT>No</ISINPUTCREDIT>
                    <ISEXEMPTED>No</ISEXEMPTED>
                    <ISABATEMENTAPPLICABLE>No</ISABATEMENTAPPLICABLE>
                    <ISSTXPARTY>No</ISSTXPARTY>
                    <ISSTXNONREALIZEDTYPE>No</ISSTXNONREALIZEDTYPE>
                    <ISUSEDFORCVD>No</ISUSEDFORCVD>
                    <LEDBELONGSTONONTAXABLE>No</LEDBELONGSTONONTAXABLE>
                    <ISEXCISEMERCHANTEXPORTER>No</ISEXCISEMERCHANTEXPORTER>
                    <ISPARTYEXEMPTED>No</ISPARTYEXEMPTED>
                    <ISSEZPARTY>No</ISSEZPARTY>
                    <TDSDEDUCTEEISSPECIALRATE>No</TDSDEDUCTEEISSPECIALRATE>
                    <ISECHEQUESUPPORTED>No</ISECHEQUESUPPORTED>
                    <ISEDDSUPPORTED>No</ISEDDSUPPORTED>
                    <HASECHEQUEDELIVERYMODE>No</HASECHEQUEDELIVERYMODE>
                    <HASECHEQUEDELIVERYTO>No</HASECHEQUEDELIVERYTO>
                    <HASECHEQUEPRINTLOCATION>No</HASECHEQUEPRINTLOCATION>
                    <HASECHEQUEPAYABLELOCATION>No</HASECHEQUEPAYABLELOCATION>
                    <HASECHEQUEBANKLOCATION>No</HASECHEQUEBANKLOCATION>
                    <HASEDDDELIVERYMODE>No</HASEDDDELIVERYMODE>
                    <HASEDDDELIVERYTO>No</HASEDDDELIVERYTO>
                    <HASEDDPRINTLOCATION>No</HASEDDPRINTLOCATION>
                    <HASEDDPAYABLELOCATION>No</HASEDDPAYABLELOCATION>
                    <HASEDDBANKLOCATION>No</HASEDDBANKLOCATION>
                    <ISEBANKINGENABLED>No</ISEBANKINGENABLED>
                    <ISEXPORTFILEENCRYPTED>No</ISEXPORTFILEENCRYPTED>
                    <ISBATCHENABLED>No</ISBATCHENABLED>
                    <ISPRODUCTCODEBASED>No</ISPRODUCTCODEBASED>
                    <HASEDDCITY>No</HASEDDCITY>
                    <HASECHEQUECITY>No</HASECHEQUECITY>
                    <ISFILENAMEFORMATSUPPORTED>No</ISFILENAMEFORMATSUPPORTED>
                    <HASCLIENTCODE>No</HASCLIENTCODE>
                    <PAYINSISBATCHAPPLICABLE>No</PAYINSISBATCHAPPLICABLE>
                    <PAYINSISFILENUMAPP>No</PAYINSISFILENUMAPP>
                    <ISSALARYTRANSGROUPEDFORBRS>No</ISSALARYTRANSGROUPEDFORBRS>
                    <ISEBANKINGSUPPORTED>No</ISEBANKINGSUPPORTED>
                    <ISSCBUAE>No</ISSCBUAE>
                    <ISBANKSTATUSAPP>No</ISBANKSTATUSAPP>
                    <ISSALARYGROUPED>No</ISSALARYGROUPED>
                    <USEFORPURCHASETAX>No</USEFORPURCHASETAX>
                    <AUDITED>No</AUDITED>
                    <SORTPOSITION> 1000</SORTPOSITION>
                    <ALTERID> 3459</ALTERID>
                    <OPENINGBALANCE>'''+str(data.iloc[row]['OB']).strip()+'''</OPENINGBALANCE>
                    <SERVICETAXDETAILS.LIST>      </SERVICETAXDETAILS.LIST>
                    <LBTREGNDETAILS.LIST>      </LBTREGNDETAILS.LIST>
                    <VATDETAILS.LIST>      </VATDETAILS.LIST>
                    <SALESTAXCESSDETAILS.LIST>      </SALESTAXCESSDETAILS.LIST>
                    <GSTDETAILS.LIST>      </GSTDETAILS.LIST>
                    <LANGUAGENAME.LIST>
                    <NAME.LIST TYPE="String">
                        <NAME>'''+str(data.iloc[row]['Ledger_Name']).strip()+'''</NAME>
                        <NAME>'''+str(data.iloc[row]['Alias']).strip()+'''</NAME>
                    </NAME.LIST>
                    <LANGUAGEID> 1033</LANGUAGEID>
                    </LANGUAGENAME.LIST>
                    <XBRLDETAIL.LIST>      </XBRLDETAIL.LIST>
                    <AUDITDETAILS.LIST>      </AUDITDETAILS.LIST>
                    <SCHVIDETAILS.LIST>      </SCHVIDETAILS.LIST>
                    <EXCISETARIFFDETAILS.LIST>      </EXCISETARIFFDETAILS.LIST>
                    <TCSCATEGORYDETAILS.LIST>      </TCSCATEGORYDETAILS.LIST>
                    <TDSCATEGORYDETAILS.LIST>      </TDSCATEGORYDETAILS.LIST>
                    <SLABPERIOD.LIST>      </SLABPERIOD.LIST>
                    <GRATUITYPERIOD.LIST>      </GRATUITYPERIOD.LIST>
                    <ADDITIONALCOMPUTATIONS.LIST>      </ADDITIONALCOMPUTATIONS.LIST>
                    <EXCISEJURISDICTIONDETAILS.LIST>      </EXCISEJURISDICTIONDETAILS.LIST>
                    <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                    <BANKALLOCATIONS.LIST>      </BANKALLOCATIONS.LIST>
                    <PAYMENTDETAILS.LIST>      </PAYMENTDETAILS.LIST>
                    <BANKEXPORTFORMATS.LIST>      </BANKEXPORTFORMATS.LIST>
                    <BILLALLOCATIONS.LIST>      </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>      </INTERESTCOLLECTION.LIST>
                    <LEDGERCLOSINGVALUES.LIST>      </LEDGERCLOSINGVALUES.LIST>
                    <LEDGERAUDITCLASS.LIST>      </LEDGERAUDITCLASS.LIST>
                    <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                    <TDSEXEMPTIONRULES.LIST>      </TDSEXEMPTIONRULES.LIST>
                    <DEDUCTINSAMEVCHRULES.LIST>      </DEDUCTINSAMEVCHRULES.LIST>
                    <LOWERDEDUCTION.LIST>      </LOWERDEDUCTION.LIST>
                    <STXABATEMENTDETAILS.LIST>      </STXABATEMENTDETAILS.LIST>
                    <LEDMULTIADDRESSLIST.LIST>      </LEDMULTIADDRESSLIST.LIST>
                    <STXTAXDETAILS.LIST>      </STXTAXDETAILS.LIST>
                    <CHEQUERANGE.LIST>      </CHEQUERANGE.LIST>
                    <DEFAULTVCHCHEQUEDETAILS.LIST>      </DEFAULTVCHCHEQUEDETAILS.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                    <BRSIMPORTEDINFO.LIST>      </BRSIMPORTEDINFO.LIST>
                    <AUTOBRSCONFIGS.LIST>      </AUTOBRSCONFIGS.LIST>
                    <BANKURENTRIES.LIST>      </BANKURENTRIES.LIST>
                    <DEFAULTCHEQUEDETAILS.LIST>      </DEFAULTCHEQUEDETAILS.LIST>
                    <DEFAULTOPENINGCHEQUEDETAILS.LIST>      </DEFAULTOPENINGCHEQUEDETAILS.LIST>
                    <CANCELLEDPAYALLOCATIONS.LIST>      </CANCELLEDPAYALLOCATIONS.LIST>
                    <ECHEQUEPRINTLOCATION.LIST>      </ECHEQUEPRINTLOCATION.LIST>
                    <ECHEQUEPAYABLELOCATION.LIST>      </ECHEQUEPAYABLELOCATION.LIST>
                    <EDDPRINTLOCATION.LIST>      </EDDPRINTLOCATION.LIST>
                    <EDDPAYABLELOCATION.LIST>      </EDDPAYABLELOCATION.LIST>
                    <AVAILABLETRANSACTIONTYPES.LIST>      </AVAILABLETRANSACTIONTYPES.LIST>
                    <LEDPAYINSCONFIGS.LIST>      </LEDPAYINSCONFIGS.LIST>
                    <TYPECODEDETAILS.LIST>      </TYPECODEDETAILS.LIST>
                    <FIELDVALIDATIONDETAILS.LIST>      </FIELDVALIDATIONDETAILS.LIST>
                    <INPUTCRALLOCS.LIST>      </INPUTCRALLOCS.LIST>
                    <GSTCLASSFNIGSTRATES.LIST>      </GSTCLASSFNIGSTRATES.LIST>
                    <EXTARIFFDUTYHEADDETAILS.LIST>      </EXTARIFFDUTYHEADDETAILS.LIST>
                    <VOUCHERTYPEPRODUCTCODES.LIST>      </VOUCHERTYPEPRODUCTCODES.LIST>
                    </LEDGER>
                </TALLYMESSAGE>
                '''

            xml_final=xml_begin+xml+xml_end
            request.session['xml_final'] = xml_final
            csv_string = data.to_csv(index=False)
            request.session['data_csv'] = csv_string
            request.session['file_name'] = 'Master_Ledger'
            # download_url = reverse('view:download_xml') + f'?xml_final={xml_final}&csv_string={csv_string}'

            
        # You can add further processing of the uploaded file here

        return render(request, 'GSTapp/master_ledger.html', {'message':True,'file_name':uploaded_file})

    
class Master_Duties(View):
    def get(self,request):
        return render(request,'GSTapp/master_duties.html')
    
    def post(self,request):
        uploaded_file = request.FILES.get('upload_file_master_duties')

        if uploaded_file:
            data=pd.read_excel(uploaded_file,sheet_name='TEMPLATE')
   
            data=data.fillna('')

            data['DATE']= datetime.today().strftime('%Y%m%d')

            data.to_excel("output_duties.xlsx",sheet_name='TEMPLATE') 

            def replace_amp(text):
                return text.replace('&', '&amp;')
            
            data['Ledger_Name'] = data['Ledger_Name'].apply(replace_amp)
            data['Group_Name'] = data['Group_Name'].apply(replace_amp)            

            xml_begin= '''
                <ENVELOPE>
                <HEADER>
                <TALLYREQUEST>Import Data</TALLYREQUEST>
                </HEADER>
                <BODY>
                <IMPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>All Masters</REPORTNAME>
                    <STATICVARIABLES>
                    <SVCURRENTCOMPANY>Demo Company</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
                <REQUESTDATA>
                '''
            xml_end='''
                </REQUESTDATA>
                </IMPORTDATA>
                </BODY>
                </ENVELOPE>
                '''
            xml=''

            for row in range(len(data)):
                xml=xml+'''
                <TALLYMESSAGE xmlns:UDF="TallyUDF">
                    <LEDGER NAME="'''+str(data.iloc[row]['Ledger_Name']).strip()+'''" RESERVEDNAME="">
                    <OLDAUDITENTRYIDS.LIST TYPE="Number">
                    <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                    </OLDAUDITENTRYIDS.LIST>
                    <CREATEDDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</CREATEDDATE>
                    <GUID>5146d6da-e8bc-454f-8fd1-116f5719b3cc-0000035d</GUID>
                    <CURRENCYNAME>₹</CURRENCYNAME>
                    <PARENT>'''+str(data.iloc[row]['Group_Name']).strip()+'''</PARENT>
                    <CREATEDBY>amit</CREATEDBY>
                    <TAXCLASSIFICATIONNAME/>
                    <TAXTYPE>GST</TAXTYPE>
                    <GSTTYPE/>
                    <APPROPRIATEFOR/>
                    <GSTDUTYHEAD>'''+str(data.iloc[row]['Tax_Type']).strip()+'''</GSTDUTYHEAD>
                    <ROUNDINGMETHOD/>
                    <SERVICECATEGORY>&#4; Not Applicable</SERVICECATEGORY>
                    <EXCISELEDGERCLASSIFICATION/>
                    <EXCISEDUTYTYPE/>
                    <EXCISENATUREOFPURCHASE/>
                    <LEDGERFBTCATEGORY/>
                    <ISBILLWISEON>No</ISBILLWISEON>
                    <ISCOSTCENTRESON>No</ISCOSTCENTRESON>
                    <ISINTERESTON>No</ISINTERESTON>
                    <ALLOWINMOBILE>No</ALLOWINMOBILE>
                    <ISCOSTTRACKINGON>No</ISCOSTTRACKINGON>
                    <ISBENEFICIARYCODEON>No</ISBENEFICIARYCODEON>
                    <PLASINCOMEEXPENSE>No</PLASINCOMEEXPENSE>
                    <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
                    <ASORIGINAL>Yes</ASORIGINAL>
                    <ISCONDENSED>No</ISCONDENSED>
                    <AFFECTSSTOCK>No</AFFECTSSTOCK>
                    <ISRATEINCLUSIVEVAT>No</ISRATEINCLUSIVEVAT>
                    <FORPAYROLL>No</FORPAYROLL>
                    <ISABCENABLED>No</ISABCENABLED>
                    <ISCREDITDAYSCHKON>No</ISCREDITDAYSCHKON>
                    <INTERESTONBILLWISE>No</INTERESTONBILLWISE>
                    <OVERRIDEINTEREST>No</OVERRIDEINTEREST>
                    <OVERRIDEADVINTEREST>No</OVERRIDEADVINTEREST>
                    <USEFORVAT>No</USEFORVAT>
                    <IGNORETDSEXEMPT>No</IGNORETDSEXEMPT>
                    <ISTCSAPPLICABLE>No</ISTCSAPPLICABLE>
                    <ISTDSAPPLICABLE>No</ISTDSAPPLICABLE>
                    <ISFBTAPPLICABLE>No</ISFBTAPPLICABLE>
                    <ISGSTAPPLICABLE>No</ISGSTAPPLICABLE>
                    <ISEXCISEAPPLICABLE>No</ISEXCISEAPPLICABLE>
                    <ISTDSEXPENSE>No</ISTDSEXPENSE>
                    <ISEDLIAPPLICABLE>No</ISEDLIAPPLICABLE>
                    <ISRELATEDPARTY>No</ISRELATEDPARTY>
                    <USEFORESIELIGIBILITY>No</USEFORESIELIGIBILITY>
                    <ISINTERESTINCLLASTDAY>No</ISINTERESTINCLLASTDAY>
                    <APPROPRIATETAXVALUE>No</APPROPRIATETAXVALUE>
                    <ISBEHAVEASDUTY>No</ISBEHAVEASDUTY>
                    <INTERESTINCLDAYOFADDITION>No</INTERESTINCLDAYOFADDITION>
                    <INTERESTINCLDAYOFDEDUCTION>No</INTERESTINCLDAYOFDEDUCTION>
                    <ISOTHTERRITORYASSESSEE>No</ISOTHTERRITORYASSESSEE>
                    <OVERRIDECREDITLIMIT>No</OVERRIDECREDITLIMIT>
                    <ISAGAINSTFORMC>No</ISAGAINSTFORMC>
                    <ISCHEQUEPRINTINGENABLED>Yes</ISCHEQUEPRINTINGENABLED>
                    <ISPAYUPLOAD>No</ISPAYUPLOAD>
                    <ISPAYBATCHONLYSAL>No</ISPAYBATCHONLYSAL>
                    <ISBNFCODESUPPORTED>No</ISBNFCODESUPPORTED>
                    <ALLOWEXPORTWITHERRORS>No</ALLOWEXPORTWITHERRORS>
                    <CONSIDERPURCHASEFOREXPORT>No</CONSIDERPURCHASEFOREXPORT>
                    <ISTRANSPORTER>No</ISTRANSPORTER>
                    <USEFORNOTIONALITC>No</USEFORNOTIONALITC>
                    <ISECOMMOPERATOR>No</ISECOMMOPERATOR>
                    <SHOWINPAYSLIP>No</SHOWINPAYSLIP>
                    <USEFORGRATUITY>No</USEFORGRATUITY>
                    <ISTDSPROJECTED>No</ISTDSPROJECTED>
                    <FORSERVICETAX>No</FORSERVICETAX>
                    <ISINPUTCREDIT>No</ISINPUTCREDIT>
                    <ISEXEMPTED>No</ISEXEMPTED>
                    <ISABATEMENTAPPLICABLE>No</ISABATEMENTAPPLICABLE>
                    <ISSTXPARTY>No</ISSTXPARTY>
                    <ISSTXNONREALIZEDTYPE>No</ISSTXNONREALIZEDTYPE>
                    <ISUSEDFORCVD>No</ISUSEDFORCVD>
                    <LEDBELONGSTONONTAXABLE>No</LEDBELONGSTONONTAXABLE>
                    <ISEXCISEMERCHANTEXPORTER>No</ISEXCISEMERCHANTEXPORTER>
                    <ISPARTYEXEMPTED>No</ISPARTYEXEMPTED>
                    <ISSEZPARTY>No</ISSEZPARTY>
                    <TDSDEDUCTEEISSPECIALRATE>No</TDSDEDUCTEEISSPECIALRATE>
                    <ISECHEQUESUPPORTED>No</ISECHEQUESUPPORTED>
                    <ISEDDSUPPORTED>No</ISEDDSUPPORTED>
                    <HASECHEQUEDELIVERYMODE>No</HASECHEQUEDELIVERYMODE>
                    <HASECHEQUEDELIVERYTO>No</HASECHEQUEDELIVERYTO>
                    <HASECHEQUEPRINTLOCATION>No</HASECHEQUEPRINTLOCATION>
                    <HASECHEQUEPAYABLELOCATION>No</HASECHEQUEPAYABLELOCATION>
                    <HASECHEQUEBANKLOCATION>No</HASECHEQUEBANKLOCATION>
                    <HASEDDDELIVERYMODE>No</HASEDDDELIVERYMODE>
                    <HASEDDDELIVERYTO>No</HASEDDDELIVERYTO>
                    <HASEDDPRINTLOCATION>No</HASEDDPRINTLOCATION>
                    <HASEDDPAYABLELOCATION>No</HASEDDPAYABLELOCATION>
                    <HASEDDBANKLOCATION>No</HASEDDBANKLOCATION>
                    <ISEBANKINGENABLED>No</ISEBANKINGENABLED>
                    <ISEXPORTFILEENCRYPTED>No</ISEXPORTFILEENCRYPTED>
                    <ISBATCHENABLED>No</ISBATCHENABLED>
                    <ISPRODUCTCODEBASED>No</ISPRODUCTCODEBASED>
                    <HASEDDCITY>No</HASEDDCITY>
                    <HASECHEQUECITY>No</HASECHEQUECITY>
                    <ISFILENAMEFORMATSUPPORTED>No</ISFILENAMEFORMATSUPPORTED>
                    <HASCLIENTCODE>No</HASCLIENTCODE>
                    <PAYINSISBATCHAPPLICABLE>No</PAYINSISBATCHAPPLICABLE>
                    <PAYINSISFILENUMAPP>No</PAYINSISFILENUMAPP>
                    <ISSALARYTRANSGROUPEDFORBRS>No</ISSALARYTRANSGROUPEDFORBRS>
                    <ISEBANKINGSUPPORTED>No</ISEBANKINGSUPPORTED>
                    <ISSCBUAE>No</ISSCBUAE>
                    <ISBANKSTATUSAPP>No</ISBANKSTATUSAPP>
                    <ISSALARYGROUPED>No</ISSALARYGROUPED>
                    <USEFORPURCHASETAX>No</USEFORPURCHASETAX>
                    <AUDITED>No</AUDITED>
                    <SORTPOSITION> 1000</SORTPOSITION>
                    <ALTERID> 3300</ALTERID>
                    <RATEOFTAXCALCULATION>'''+(str(int(data.iloc[row]['Rate_of_Tax']) if (isinstance(data.iloc[row]['Rate_of_Tax'], (int, float)) and (data.iloc[row]['Rate_of_Tax']>3)) else f"{data.iloc[row]['Rate_of_Tax']:.2f}")).strip()+'''</RATEOFTAXCALCULATION>
                    <SERVICETAXDETAILS.LIST>      </SERVICETAXDETAILS.LIST>
                    <LBTREGNDETAILS.LIST>      </LBTREGNDETAILS.LIST>
                    <VATDETAILS.LIST>      </VATDETAILS.LIST>
                    <SALESTAXCESSDETAILS.LIST>      </SALESTAXCESSDETAILS.LIST>
                    <GSTDETAILS.LIST>      </GSTDETAILS.LIST>
                    <LANGUAGENAME.LIST>
                    <NAME.LIST TYPE="String">
                        <NAME>'''+str(data.iloc[row]['Ledger_Name']).strip()+'''</NAME>
                    </NAME.LIST>
                    <LANGUAGEID> 1033</LANGUAGEID>
                    </LANGUAGENAME.LIST>
                    <XBRLDETAIL.LIST>      </XBRLDETAIL.LIST>
                    <AUDITDETAILS.LIST>      </AUDITDETAILS.LIST>
                    <SCHVIDETAILS.LIST>      </SCHVIDETAILS.LIST>
                    <EXCISETARIFFDETAILS.LIST>      </EXCISETARIFFDETAILS.LIST>
                    <TCSCATEGORYDETAILS.LIST>      </TCSCATEGORYDETAILS.LIST>
                    <TDSCATEGORYDETAILS.LIST>      </TDSCATEGORYDETAILS.LIST>
                    <SLABPERIOD.LIST>      </SLABPERIOD.LIST>
                    <GRATUITYPERIOD.LIST>      </GRATUITYPERIOD.LIST>
                    <ADDITIONALCOMPUTATIONS.LIST>      </ADDITIONALCOMPUTATIONS.LIST>
                    <EXCISEJURISDICTIONDETAILS.LIST>      </EXCISEJURISDICTIONDETAILS.LIST>
                    <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                    <BANKALLOCATIONS.LIST>      </BANKALLOCATIONS.LIST>
                    <PAYMENTDETAILS.LIST>      </PAYMENTDETAILS.LIST>
                    <BANKEXPORTFORMATS.LIST>      </BANKEXPORTFORMATS.LIST>
                    <BILLALLOCATIONS.LIST>      </BILLALLOCATIONS.LIST>
                    <INTERESTCOLLECTION.LIST>      </INTERESTCOLLECTION.LIST>
                    <LEDGERCLOSINGVALUES.LIST>      </LEDGERCLOSINGVALUES.LIST>
                    <LEDGERAUDITCLASS.LIST>      </LEDGERAUDITCLASS.LIST>
                    <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                    <TDSEXEMPTIONRULES.LIST>      </TDSEXEMPTIONRULES.LIST>
                    <DEDUCTINSAMEVCHRULES.LIST>      </DEDUCTINSAMEVCHRULES.LIST>
                    <LOWERDEDUCTION.LIST>      </LOWERDEDUCTION.LIST>
                    <STXABATEMENTDETAILS.LIST>      </STXABATEMENTDETAILS.LIST>
                    <LEDMULTIADDRESSLIST.LIST>      </LEDMULTIADDRESSLIST.LIST>
                    <STXTAXDETAILS.LIST>      </STXTAXDETAILS.LIST>
                    <CHEQUERANGE.LIST>      </CHEQUERANGE.LIST>
                    <DEFAULTVCHCHEQUEDETAILS.LIST>      </DEFAULTVCHCHEQUEDETAILS.LIST>
                    <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                    <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                    <BRSIMPORTEDINFO.LIST>      </BRSIMPORTEDINFO.LIST>
                    <AUTOBRSCONFIGS.LIST>      </AUTOBRSCONFIGS.LIST>
                    <BANKURENTRIES.LIST>      </BANKURENTRIES.LIST>
                    <DEFAULTCHEQUEDETAILS.LIST>      </DEFAULTCHEQUEDETAILS.LIST>
                    <DEFAULTOPENINGCHEQUEDETAILS.LIST>      </DEFAULTOPENINGCHEQUEDETAILS.LIST>
                    <CANCELLEDPAYALLOCATIONS.LIST>      </CANCELLEDPAYALLOCATIONS.LIST>
                    <ECHEQUEPRINTLOCATION.LIST>      </ECHEQUEPRINTLOCATION.LIST>
                    <ECHEQUEPAYABLELOCATION.LIST>      </ECHEQUEPAYABLELOCATION.LIST>
                    <EDDPRINTLOCATION.LIST>      </EDDPRINTLOCATION.LIST>
                    <EDDPAYABLELOCATION.LIST>      </EDDPAYABLELOCATION.LIST>
                    <AVAILABLETRANSACTIONTYPES.LIST>      </AVAILABLETRANSACTIONTYPES.LIST>
                    <LEDPAYINSCONFIGS.LIST>      </LEDPAYINSCONFIGS.LIST>
                    <TYPECODEDETAILS.LIST>      </TYPECODEDETAILS.LIST>
                    <FIELDVALIDATIONDETAILS.LIST>      </FIELDVALIDATIONDETAILS.LIST>
                    <INPUTCRALLOCS.LIST>      </INPUTCRALLOCS.LIST>
                    <GSTCLASSFNIGSTRATES.LIST>      </GSTCLASSFNIGSTRATES.LIST>
                    <EXTARIFFDUTYHEADDETAILS.LIST>      </EXTARIFFDUTYHEADDETAILS.LIST>
                    <VOUCHERTYPEPRODUCTCODES.LIST>      </VOUCHERTYPEPRODUCTCODES.LIST>
                    </LEDGER>
                </TALLYMESSAGE>
                '''

            xml_final=xml_begin+xml+xml_end
            request.session['xml_final'] = xml_final
            csv_string = data.to_csv(index=False)
            request.session['data_csv'] = csv_string
            request.session['file_name'] = 'Master_Duties'
            # download_url = reverse('view:download_xml') + f'?xml_final={xml_final}&csv_string={csv_string}'

            
        # You can add further processing of the uploaded file here

        return render(request, 'GSTapp/master_ledger.html', {'message':True,'file_name':uploaded_file})

class Master_PS(View):
    def get(self,request):
        return render(request,'GSTapp/master_ps.html')
    
    def post(self,request):
        uploaded_file = request.FILES.get('upload_file_master_ps')

        if uploaded_file:
            data=pd.read_excel(uploaded_file,sheet_name='TEMPLATE')
   
            data=data.fillna('')

            data['DATE']= datetime.today().strftime('%Y%m%d')

            data.to_excel("output_ps.xlsx",sheet_name='TEMPLATE') 

            def replace_amp(text):
                return text.replace('&', '&amp;')
            
            data['Ledger_Name'] = data['Ledger_Name'].apply(replace_amp)
            data['Group_Name'] = data['Group_Name'].apply(replace_amp) 
            data['Nature_of_transaction'] = data['Nature_of_transaction'].apply(replace_amp)                        

            xml_begin= '''
                <ENVELOPE>
                <HEADER>
                <TALLYREQUEST>Import Data</TALLYREQUEST>
                </HEADER>
                <BODY>
                <IMPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>All Masters</REPORTNAME>
                    <STATICVARIABLES>
                    <SVCURRENTCOMPANY>Demo Company</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
                <REQUESTDATA>
                '''
            xml_end='''
                </REQUESTDATA>
                </IMPORTDATA>
                </BODY>
                </ENVELOPE>
                '''
            xml=''

            for row in range(len(data)):
                if (('Exempt' in data.iloc[row]['Nature_of_transaction']) or (data.iloc[row]['RATE_OF_CGST_SGST']==0) or (data.iloc[row]['RATE_OF_IGST']==0)):    
                    xml=xml+'''
                    <TALLYMESSAGE xmlns:UDF="TallyUDF">
                        <LEDGER NAME="'''+str(data.iloc[row]['Ledger_Name']).strip()+'''" RESERVEDNAME="">
                        <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                        </OLDAUDITENTRYIDS.LIST>
                        <CREATEDDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</CREATEDDATE>
                                                        
                        <GUID>5146d6da-e8bc-454f-8fd1-116f5719b3cc-0000039c</GUID>
                        <CURRENCYNAME>₹</CURRENCYNAME>
                        <PARENT>'''+str(data.iloc[row]['Group_Name']).strip()+'''</PARENT>
                        <GSTAPPLICABLE>&#4; Applicable</GSTAPPLICABLE>
                        <CREATEDBY>amit</CREATEDBY>
                                                    
                        <TAXCLASSIFICATIONNAME/>
                        <TAXTYPE>Others</TAXTYPE>
                        <LEDADDLALLOCTYPE/>
                        <GSTTYPE/>
                        <APPROPRIATEFOR/>
                        <GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>
                        <SERVICECATEGORY>&#4; Not Applicable</SERVICECATEGORY>
                        <EXCISELEDGERCLASSIFICATION/>
                        <EXCISEDUTYTYPE/>
                        <EXCISENATUREOFPURCHASE/>
                        <LEDGERFBTCATEGORY/>
                        <VATAPPLICABLE>&#4; Applicable</VATAPPLICABLE>
                        <ISBILLWISEON>No</ISBILLWISEON>
                        <ISCOSTCENTRESON>Yes</ISCOSTCENTRESON>
                        <ISINTERESTON>No</ISINTERESTON>
                        <ALLOWINMOBILE>No</ALLOWINMOBILE>
                        <ISCOSTTRACKINGON>No</ISCOSTTRACKINGON>
                        <ISBENEFICIARYCODEON>No</ISBENEFICIARYCODEON>
                        <PLASINCOMEEXPENSE>No</PLASINCOMEEXPENSE>
                        <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
                        <ASORIGINAL>Yes</ASORIGINAL>
                        <ISCONDENSED>No</ISCONDENSED>
                        <AFFECTSSTOCK>No</AFFECTSSTOCK>
                        <ISRATEINCLUSIVEVAT>No</ISRATEINCLUSIVEVAT>
                        <FORPAYROLL>No</FORPAYROLL>
                        <ISABCENABLED>No</ISABCENABLED>
                        <ISCREDITDAYSCHKON>No</ISCREDITDAYSCHKON>
                        <INTERESTONBILLWISE>No</INTERESTONBILLWISE>
                        <OVERRIDEINTEREST>No</OVERRIDEINTEREST>
                        <OVERRIDEADVINTEREST>No</OVERRIDEADVINTEREST>
                        <USEFORVAT>No</USEFORVAT>
                        <IGNORETDSEXEMPT>No</IGNORETDSEXEMPT>
                        <ISTCSAPPLICABLE>No</ISTCSAPPLICABLE>
                        <ISTDSAPPLICABLE>No</ISTDSAPPLICABLE>
                        <ISFBTAPPLICABLE>No</ISFBTAPPLICABLE>
                        <ISGSTAPPLICABLE>No</ISGSTAPPLICABLE>
                        <ISEXCISEAPPLICABLE>No</ISEXCISEAPPLICABLE>
                        <ISTDSEXPENSE>No</ISTDSEXPENSE>
                        <ISEDLIAPPLICABLE>No</ISEDLIAPPLICABLE>
                        <ISRELATEDPARTY>No</ISRELATEDPARTY>
                        <USEFORESIELIGIBILITY>No</USEFORESIELIGIBILITY>
                        <ISINTERESTINCLLASTDAY>No</ISINTERESTINCLLASTDAY>
                        <APPROPRIATETAXVALUE>No</APPROPRIATETAXVALUE>
                        <ISBEHAVEASDUTY>No</ISBEHAVEASDUTY>
                        <INTERESTINCLDAYOFADDITION>No</INTERESTINCLDAYOFADDITION>
                        <INTERESTINCLDAYOFDEDUCTION>No</INTERESTINCLDAYOFDEDUCTION>
                        <ISOTHTERRITORYASSESSEE>No</ISOTHTERRITORYASSESSEE>
                        <OVERRIDECREDITLIMIT>No</OVERRIDECREDITLIMIT>
                        <ISAGAINSTFORMC>No</ISAGAINSTFORMC>
                        <ISCHEQUEPRINTINGENABLED>Yes</ISCHEQUEPRINTINGENABLED>
                        <ISPAYUPLOAD>No</ISPAYUPLOAD>
                        <ISPAYBATCHONLYSAL>No</ISPAYBATCHONLYSAL>
                        <ISBNFCODESUPPORTED>No</ISBNFCODESUPPORTED>
                        <ALLOWEXPORTWITHERRORS>No</ALLOWEXPORTWITHERRORS>
                        <CONSIDERPURCHASEFOREXPORT>No</CONSIDERPURCHASEFOREXPORT>
                        <ISTRANSPORTER>No</ISTRANSPORTER>
                        <USEFORNOTIONALITC>No</USEFORNOTIONALITC>
                        <ISECOMMOPERATOR>No</ISECOMMOPERATOR>
                        <SHOWINPAYSLIP>No</SHOWINPAYSLIP>
                        <USEFORGRATUITY>No</USEFORGRATUITY>
                        <ISTDSPROJECTED>No</ISTDSPROJECTED>
                        <FORSERVICETAX>No</FORSERVICETAX>
                        <ISINPUTCREDIT>No</ISINPUTCREDIT>
                        <ISEXEMPTED>No</ISEXEMPTED>
                        <ISABATEMENTAPPLICABLE>No</ISABATEMENTAPPLICABLE>
                        <ISSTXPARTY>No</ISSTXPARTY>
                        <ISSTXNONREALIZEDTYPE>No</ISSTXNONREALIZEDTYPE>
                        <ISUSEDFORCVD>No</ISUSEDFORCVD>
                        <LEDBELONGSTONONTAXABLE>No</LEDBELONGSTONONTAXABLE>
                        <ISEXCISEMERCHANTEXPORTER>No</ISEXCISEMERCHANTEXPORTER>
                        <ISPARTYEXEMPTED>No</ISPARTYEXEMPTED>
                        <ISSEZPARTY>No</ISSEZPARTY>
                        <TDSDEDUCTEEISSPECIALRATE>No</TDSDEDUCTEEISSPECIALRATE>
                        <ISECHEQUESUPPORTED>No</ISECHEQUESUPPORTED>
                        <ISEDDSUPPORTED>No</ISEDDSUPPORTED>
                        <HASECHEQUEDELIVERYMODE>No</HASECHEQUEDELIVERYMODE>
                        <HASECHEQUEDELIVERYTO>No</HASECHEQUEDELIVERYTO>
                        <HASECHEQUEPRINTLOCATION>No</HASECHEQUEPRINTLOCATION>
                        <HASECHEQUEPAYABLELOCATION>No</HASECHEQUEPAYABLELOCATION>
                        <HASECHEQUEBANKLOCATION>No</HASECHEQUEBANKLOCATION>
                        <HASEDDDELIVERYMODE>No</HASEDDDELIVERYMODE>
                        <HASEDDDELIVERYTO>No</HASEDDDELIVERYTO>
                        <HASEDDPRINTLOCATION>No</HASEDDPRINTLOCATION>
                        <HASEDDPAYABLELOCATION>No</HASEDDPAYABLELOCATION>
                        <HASEDDBANKLOCATION>No</HASEDDBANKLOCATION>
                        <ISEBANKINGENABLED>No</ISEBANKINGENABLED>
                        <ISEXPORTFILEENCRYPTED>No</ISEXPORTFILEENCRYPTED>
                        <ISBATCHENABLED>No</ISBATCHENABLED>
                        <ISPRODUCTCODEBASED>No</ISPRODUCTCODEBASED>
                        <HASEDDCITY>No</HASEDDCITY>
                        <HASECHEQUECITY>No</HASECHEQUECITY>
                        <ISFILENAMEFORMATSUPPORTED>No</ISFILENAMEFORMATSUPPORTED>
                        <HASCLIENTCODE>No</HASCLIENTCODE>
                        <PAYINSISBATCHAPPLICABLE>No</PAYINSISBATCHAPPLICABLE>
                        <PAYINSISFILENUMAPP>No</PAYINSISFILENUMAPP>
                        <ISSALARYTRANSGROUPEDFORBRS>No</ISSALARYTRANSGROUPEDFORBRS>
                        <ISEBANKINGSUPPORTED>No</ISEBANKINGSUPPORTED>
                        <ISSCBUAE>No</ISSCBUAE>
                        <ISBANKSTATUSAPP>No</ISBANKSTATUSAPP>
                        <ISSALARYGROUPED>No</ISSALARYGROUPED>
                        <USEFORPURCHASETAX>No</USEFORPURCHASETAX>
                        <AUDITED>No</AUDITED>
                        <SORTPOSITION> 1000</SORTPOSITION>
                        <ALTERID> 3467</ALTERID>
                        <SERVICETAXDETAILS.LIST>      </SERVICETAXDETAILS.LIST>
                        <LBTREGNDETAILS.LIST>      </LBTREGNDETAILS.LIST>
                        <VATDETAILS.LIST>      </VATDETAILS.LIST>
                        <SALESTAXCESSDETAILS.LIST>      </SALESTAXCESSDETAILS.LIST>
                        <GSTDETAILS.LIST>
                        <APPLICABLEFROM>20100101</APPLICABLEFROM>
                        <HSNMASTERNAME/>
                        <TAXABILITY>Exempt</TAXABILITY>
                        <GSTNATUREOFTRANSACTION>'''+str(data.iloc[row]['Nature_of_transaction']).strip()+'''</GSTNATUREOFTRANSACTION>
                        <ISREVERSECHARGEAPPLICABLE>No</ISREVERSECHARGEAPPLICABLE>
                        <ISNONGSTGOODS>No</ISNONGSTGOODS>
                        <GSTINELIGIBLEITC>No</GSTINELIGIBLEITC>
                        <INCLUDEEXPFORSLABCALC>No</INCLUDEEXPFORSLABCALC>
                        <STATEWISEDETAILS.LIST>
                            <STATENAME>&#4; Any</STATENAME>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Central Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                                                
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>State Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                                                
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Integrated Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                                                
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Cess</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Cess on Qty</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Quantity</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>State Cess</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <GSTSLABRATES.LIST>        </GSTSLABRATES.LIST>
                        </STATEWISEDETAILS.LIST>
                        <TEMPGSTDETAILSLABRATES.LIST>       </TEMPGSTDETAILSLABRATES.LIST>
                        </GSTDETAILS.LIST>
                        <LANGUAGENAME.LIST>
                        <NAME.LIST TYPE="String">
                            <NAME>'''+str(data.iloc[row]['Ledger_Name']).strip()+'''</NAME>
                        </NAME.LIST>
                        <LANGUAGEID> 1033</LANGUAGEID>
                        </LANGUAGENAME.LIST>
                        <XBRLDETAIL.LIST>      </XBRLDETAIL.LIST>
                        <AUDITDETAILS.LIST>      </AUDITDETAILS.LIST>
                        <SCHVIDETAILS.LIST>      </SCHVIDETAILS.LIST>
                        <EXCISETARIFFDETAILS.LIST>      </EXCISETARIFFDETAILS.LIST>
                        <TCSCATEGORYDETAILS.LIST>      </TCSCATEGORYDETAILS.LIST>
                        <TDSCATEGORYDETAILS.LIST>      </TDSCATEGORYDETAILS.LIST>
                        <SLABPERIOD.LIST>      </SLABPERIOD.LIST>
                        <GRATUITYPERIOD.LIST>      </GRATUITYPERIOD.LIST>
                        <ADDITIONALCOMPUTATIONS.LIST>      </ADDITIONALCOMPUTATIONS.LIST>
                        <EXCISEJURISDICTIONDETAILS.LIST>      </EXCISEJURISDICTIONDETAILS.LIST>
                        <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                        <BANKALLOCATIONS.LIST>      </BANKALLOCATIONS.LIST>
                        <PAYMENTDETAILS.LIST>      </PAYMENTDETAILS.LIST>
                        <BANKEXPORTFORMATS.LIST>      </BANKEXPORTFORMATS.LIST>
                        <BILLALLOCATIONS.LIST>      </BILLALLOCATIONS.LIST>
                        <INTERESTCOLLECTION.LIST>      </INTERESTCOLLECTION.LIST>
                        <LEDGERCLOSINGVALUES.LIST>      </LEDGERCLOSINGVALUES.LIST>
                        <LEDGERAUDITCLASS.LIST>      </LEDGERAUDITCLASS.LIST>
                        <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                        <TDSEXEMPTIONRULES.LIST>      </TDSEXEMPTIONRULES.LIST>
                        <DEDUCTINSAMEVCHRULES.LIST>      </DEDUCTINSAMEVCHRULES.LIST>
                        <LOWERDEDUCTION.LIST>      </LOWERDEDUCTION.LIST>
                        <STXABATEMENTDETAILS.LIST>      </STXABATEMENTDETAILS.LIST>
                        <LEDMULTIADDRESSLIST.LIST>      </LEDMULTIADDRESSLIST.LIST>
                        <STXTAXDETAILS.LIST>      </STXTAXDETAILS.LIST>
                        <CHEQUERANGE.LIST>      </CHEQUERANGE.LIST>
                        <DEFAULTVCHCHEQUEDETAILS.LIST>      </DEFAULTVCHCHEQUEDETAILS.LIST>
                        <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                        <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                        <BRSIMPORTEDINFO.LIST>      </BRSIMPORTEDINFO.LIST>
                        <AUTOBRSCONFIGS.LIST>      </AUTOBRSCONFIGS.LIST>
                        <BANKURENTRIES.LIST>      </BANKURENTRIES.LIST>
                        <DEFAULTCHEQUEDETAILS.LIST>      </DEFAULTCHEQUEDETAILS.LIST>
                        <DEFAULTOPENINGCHEQUEDETAILS.LIST>      </DEFAULTOPENINGCHEQUEDETAILS.LIST>
                        <CANCELLEDPAYALLOCATIONS.LIST>      </CANCELLEDPAYALLOCATIONS.LIST>
                        <ECHEQUEPRINTLOCATION.LIST>      </ECHEQUEPRINTLOCATION.LIST>
                        <ECHEQUEPAYABLELOCATION.LIST>      </ECHEQUEPAYABLELOCATION.LIST>
                        <EDDPRINTLOCATION.LIST>      </EDDPRINTLOCATION.LIST>
                        <EDDPAYABLELOCATION.LIST>      </EDDPAYABLELOCATION.LIST>
                        <AVAILABLETRANSACTIONTYPES.LIST>      </AVAILABLETRANSACTIONTYPES.LIST>
                        <LEDPAYINSCONFIGS.LIST>      </LEDPAYINSCONFIGS.LIST>
                        <TYPECODEDETAILS.LIST>      </TYPECODEDETAILS.LIST>
                        <FIELDVALIDATIONDETAILS.LIST>      </FIELDVALIDATIONDETAILS.LIST>
                        <INPUTCRALLOCS.LIST>      </INPUTCRALLOCS.LIST>
                        <GSTCLASSFNIGSTRATES.LIST>      </GSTCLASSFNIGSTRATES.LIST>
                        <EXTARIFFDUTYHEADDETAILS.LIST>      </EXTARIFFDUTYHEADDETAILS.LIST>
                        <VOUCHERTYPEPRODUCTCODES.LIST>      </VOUCHERTYPEPRODUCTCODES.LIST>
                        </LEDGER>
                    </TALLYMESSAGE>
                    '''
                else: 
                    xml=xml+'''
                    <TALLYMESSAGE xmlns:UDF="TallyUDF">
                        <LEDGER NAME="'''+str(data.iloc[row]['Ledger_Name']).strip()+'''" RESERVEDNAME="">
                        <OLDAUDITENTRYIDS.LIST TYPE="Number">
                        <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                        </OLDAUDITENTRYIDS.LIST>
                        <CREATEDDATE>'''+str(data.iloc[row]['DATE']).strip()+'''</CREATEDDATE>
                        <ALTEREDON>'''+str(data.iloc[row]['DATE']).strip()+'''</ALTEREDON>
                        <GUID>5146d6da-e8bc-454f-8fd1-116f5719b3cc-0000036c</GUID>
                        <CURRENCYNAME>₹</CURRENCYNAME>
                        <PARENT>'''+str(data.iloc[row]['Group_Name']).strip()+'''</PARENT>
                        <GSTAPPLICABLE>&#4; Applicable</GSTAPPLICABLE>
                        <CREATEDBY>amit</CREATEDBY>
                        <ALTEREDBY>amit</ALTEREDBY>
                        <TAXCLASSIFICATIONNAME/>
                        <TAXTYPE>Others</TAXTYPE>
                        <LEDADDLALLOCTYPE/>
                        <GSTTYPE/>
                        <APPROPRIATEFOR/>
                        <GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>
                        <SERVICECATEGORY>&#4; Not Applicable</SERVICECATEGORY>
                        <EXCISELEDGERCLASSIFICATION/>
                        <EXCISEDUTYTYPE/>
                        <EXCISENATUREOFPURCHASE/>
                        <LEDGERFBTCATEGORY/>
                        <VATAPPLICABLE>&#4; Applicable</VATAPPLICABLE>
                        <ISBILLWISEON>No</ISBILLWISEON>
                        <ISCOSTCENTRESON>Yes</ISCOSTCENTRESON>
                        <ISINTERESTON>No</ISINTERESTON>
                        <ALLOWINMOBILE>No</ALLOWINMOBILE>
                        <ISCOSTTRACKINGON>No</ISCOSTTRACKINGON>
                        <ISBENEFICIARYCODEON>No</ISBENEFICIARYCODEON>
                        <PLASINCOMEEXPENSE>No</PLASINCOMEEXPENSE>
                        <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
                        <ASORIGINAL>Yes</ASORIGINAL>
                        <ISCONDENSED>No</ISCONDENSED>
                        <AFFECTSSTOCK>No</AFFECTSSTOCK>
                        <ISRATEINCLUSIVEVAT>No</ISRATEINCLUSIVEVAT>
                        <FORPAYROLL>No</FORPAYROLL>
                        <ISABCENABLED>No</ISABCENABLED>
                        <ISCREDITDAYSCHKON>No</ISCREDITDAYSCHKON>
                        <INTERESTONBILLWISE>No</INTERESTONBILLWISE>
                        <OVERRIDEINTEREST>No</OVERRIDEINTEREST>
                        <OVERRIDEADVINTEREST>No</OVERRIDEADVINTEREST>
                        <USEFORVAT>No</USEFORVAT>
                        <IGNORETDSEXEMPT>No</IGNORETDSEXEMPT>
                        <ISTCSAPPLICABLE>No</ISTCSAPPLICABLE>
                        <ISTDSAPPLICABLE>No</ISTDSAPPLICABLE>
                        <ISFBTAPPLICABLE>No</ISFBTAPPLICABLE>
                        <ISGSTAPPLICABLE>No</ISGSTAPPLICABLE>
                        <ISEXCISEAPPLICABLE>No</ISEXCISEAPPLICABLE>
                        <ISTDSEXPENSE>No</ISTDSEXPENSE>
                        <ISEDLIAPPLICABLE>No</ISEDLIAPPLICABLE>
                        <ISRELATEDPARTY>No</ISRELATEDPARTY>
                        <USEFORESIELIGIBILITY>No</USEFORESIELIGIBILITY>
                        <ISINTERESTINCLLASTDAY>No</ISINTERESTINCLLASTDAY>
                        <APPROPRIATETAXVALUE>No</APPROPRIATETAXVALUE>
                        <ISBEHAVEASDUTY>No</ISBEHAVEASDUTY>
                        <INTERESTINCLDAYOFADDITION>No</INTERESTINCLDAYOFADDITION>
                        <INTERESTINCLDAYOFDEDUCTION>No</INTERESTINCLDAYOFDEDUCTION>
                        <ISOTHTERRITORYASSESSEE>No</ISOTHTERRITORYASSESSEE>
                        <OVERRIDECREDITLIMIT>No</OVERRIDECREDITLIMIT>
                        <ISAGAINSTFORMC>No</ISAGAINSTFORMC>
                        <ISCHEQUEPRINTINGENABLED>Yes</ISCHEQUEPRINTINGENABLED>
                        <ISPAYUPLOAD>No</ISPAYUPLOAD>
                        <ISPAYBATCHONLYSAL>No</ISPAYBATCHONLYSAL>
                        <ISBNFCODESUPPORTED>No</ISBNFCODESUPPORTED>
                        <ALLOWEXPORTWITHERRORS>No</ALLOWEXPORTWITHERRORS>
                        <CONSIDERPURCHASEFOREXPORT>No</CONSIDERPURCHASEFOREXPORT>
                        <ISTRANSPORTER>No</ISTRANSPORTER>
                        <USEFORNOTIONALITC>No</USEFORNOTIONALITC>
                        <ISECOMMOPERATOR>No</ISECOMMOPERATOR>
                        <SHOWINPAYSLIP>No</SHOWINPAYSLIP>
                        <USEFORGRATUITY>No</USEFORGRATUITY>
                        <ISTDSPROJECTED>No</ISTDSPROJECTED>
                        <FORSERVICETAX>No</FORSERVICETAX>
                        <ISINPUTCREDIT>No</ISINPUTCREDIT>
                        <ISEXEMPTED>No</ISEXEMPTED>
                        <ISABATEMENTAPPLICABLE>No</ISABATEMENTAPPLICABLE>
                        <ISSTXPARTY>No</ISSTXPARTY>
                        <ISSTXNONREALIZEDTYPE>No</ISSTXNONREALIZEDTYPE>
                        <ISUSEDFORCVD>No</ISUSEDFORCVD>
                        <LEDBELONGSTONONTAXABLE>No</LEDBELONGSTONONTAXABLE>
                        <ISEXCISEMERCHANTEXPORTER>No</ISEXCISEMERCHANTEXPORTER>
                        <ISPARTYEXEMPTED>No</ISPARTYEXEMPTED>
                        <ISSEZPARTY>No</ISSEZPARTY>
                        <TDSDEDUCTEEISSPECIALRATE>No</TDSDEDUCTEEISSPECIALRATE>
                        <ISECHEQUESUPPORTED>No</ISECHEQUESUPPORTED>
                        <ISEDDSUPPORTED>No</ISEDDSUPPORTED>
                        <HASECHEQUEDELIVERYMODE>No</HASECHEQUEDELIVERYMODE>
                        <HASECHEQUEDELIVERYTO>No</HASECHEQUEDELIVERYTO>
                        <HASECHEQUEPRINTLOCATION>No</HASECHEQUEPRINTLOCATION>
                        <HASECHEQUEPAYABLELOCATION>No</HASECHEQUEPAYABLELOCATION>
                        <HASECHEQUEBANKLOCATION>No</HASECHEQUEBANKLOCATION>
                        <HASEDDDELIVERYMODE>No</HASEDDDELIVERYMODE>
                        <HASEDDDELIVERYTO>No</HASEDDDELIVERYTO>
                        <HASEDDPRINTLOCATION>No</HASEDDPRINTLOCATION>
                        <HASEDDPAYABLELOCATION>No</HASEDDPAYABLELOCATION>
                        <HASEDDBANKLOCATION>No</HASEDDBANKLOCATION>
                        <ISEBANKINGENABLED>No</ISEBANKINGENABLED>
                        <ISEXPORTFILEENCRYPTED>No</ISEXPORTFILEENCRYPTED>
                        <ISBATCHENABLED>No</ISBATCHENABLED>
                        <ISPRODUCTCODEBASED>No</ISPRODUCTCODEBASED>
                        <HASEDDCITY>No</HASEDDCITY>
                        <HASECHEQUECITY>No</HASECHEQUECITY>
                        <ISFILENAMEFORMATSUPPORTED>No</ISFILENAMEFORMATSUPPORTED>
                        <HASCLIENTCODE>No</HASCLIENTCODE>
                        <PAYINSISBATCHAPPLICABLE>No</PAYINSISBATCHAPPLICABLE>
                        <PAYINSISFILENUMAPP>No</PAYINSISFILENUMAPP>
                        <ISSALARYTRANSGROUPEDFORBRS>No</ISSALARYTRANSGROUPEDFORBRS>
                        <ISEBANKINGSUPPORTED>No</ISEBANKINGSUPPORTED>
                        <ISSCBUAE>No</ISSCBUAE>
                        <ISBANKSTATUSAPP>No</ISBANKSTATUSAPP>
                        <ISSALARYGROUPED>No</ISSALARYGROUPED>
                        <USEFORPURCHASETAX>No</USEFORPURCHASETAX>
                        <AUDITED>No</AUDITED>
                        <SORTPOSITION> 1000</SORTPOSITION>
                        <ALTERID> 3372</ALTERID>
                        <SERVICETAXDETAILS.LIST>      </SERVICETAXDETAILS.LIST>
                        <LBTREGNDETAILS.LIST>      </LBTREGNDETAILS.LIST>
                        <VATDETAILS.LIST>      </VATDETAILS.LIST>
                        <SALESTAXCESSDETAILS.LIST>      </SALESTAXCESSDETAILS.LIST>
                        <GSTDETAILS.LIST>
                        <APPLICABLEFROM>20100101</APPLICABLEFROM>
                        <HSNMASTERNAME/>
                        <TAXABILITY>Taxable</TAXABILITY>
                        <GSTNATUREOFTRANSACTION>'''+str(data.iloc[row]['Nature_of_transaction']).strip()+'''</GSTNATUREOFTRANSACTION>
                        <ISREVERSECHARGEAPPLICABLE>No</ISREVERSECHARGEAPPLICABLE>
                        <ISNONGSTGOODS>No</ISNONGSTGOODS>
                        <GSTINELIGIBLEITC>No</GSTINELIGIBLEITC>
                        <INCLUDEEXPFORSLABCALC>No</INCLUDEEXPFORSLABCALC>
                        <STATEWISEDETAILS.LIST>
                            <STATENAME>&#4; Any</STATENAME>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Central Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            <GSTRATE> '''+(str(int(data.iloc[row]['RATE_OF_CGST_SGST']) if (isinstance(data.iloc[row]['RATE_OF_CGST_SGST'], (int, float)) and (data.iloc[row]['RATE_OF_CGST_SGST']>3)) else f"{data.iloc[row]['RATE_OF_CGST_SGST']:.2f}")).strip()+'''</GSTRATE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>State Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            <GSTRATE> '''+(str(int(data.iloc[row]['RATE_OF_CGST_SGST']) if (isinstance(data.iloc[row]['RATE_OF_CGST_SGST'], (int, float)) and (data.iloc[row]['RATE_OF_CGST_SGST']>3)) else f"{data.iloc[row]['RATE_OF_CGST_SGST']:.2f}")).strip()+'''</GSTRATE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Integrated Tax</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            <GSTRATE> '''+(str(int(data.iloc[row]['RATE_OF_IGST']) if (isinstance(data.iloc[row]['RATE_OF_IGST'], (int, float)) and (data.iloc[row]['RATE_OF_IGST']>3)) else f"{data.iloc[row]['RATE_OF_IGST']:.2f}")).strip()+'''</GSTRATE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Cess</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>Cess on Qty</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Quantity</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <RATEDETAILS.LIST>
                            <GSTRATEDUTYHEAD>State Cess</GSTRATEDUTYHEAD>
                            <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                            </RATEDETAILS.LIST>
                            <GSTSLABRATES.LIST>        </GSTSLABRATES.LIST>
                        </STATEWISEDETAILS.LIST>
                        <TEMPGSTDETAILSLABRATES.LIST>       </TEMPGSTDETAILSLABRATES.LIST>
                        </GSTDETAILS.LIST>
                        <LANGUAGENAME.LIST>
                        <NAME.LIST TYPE="String">
                            <NAME>'''+str(data.iloc[row]['Ledger_Name']).strip()+'''</NAME>
                        </NAME.LIST>
                        <LANGUAGEID> 1033</LANGUAGEID>
                        </LANGUAGENAME.LIST>
                        <XBRLDETAIL.LIST>      </XBRLDETAIL.LIST>
                        <AUDITDETAILS.LIST>      </AUDITDETAILS.LIST>
                        <SCHVIDETAILS.LIST>      </SCHVIDETAILS.LIST>
                        <EXCISETARIFFDETAILS.LIST>      </EXCISETARIFFDETAILS.LIST>
                        <TCSCATEGORYDETAILS.LIST>      </TCSCATEGORYDETAILS.LIST>
                        <TDSCATEGORYDETAILS.LIST>      </TDSCATEGORYDETAILS.LIST>
                        <SLABPERIOD.LIST>      </SLABPERIOD.LIST>
                        <GRATUITYPERIOD.LIST>      </GRATUITYPERIOD.LIST>
                        <ADDITIONALCOMPUTATIONS.LIST>      </ADDITIONALCOMPUTATIONS.LIST>
                        <EXCISEJURISDICTIONDETAILS.LIST>      </EXCISEJURISDICTIONDETAILS.LIST>
                        <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                        <BANKALLOCATIONS.LIST>      </BANKALLOCATIONS.LIST>
                        <PAYMENTDETAILS.LIST>      </PAYMENTDETAILS.LIST>
                        <BANKEXPORTFORMATS.LIST>      </BANKEXPORTFORMATS.LIST>
                        <BILLALLOCATIONS.LIST>      </BILLALLOCATIONS.LIST>
                        <INTERESTCOLLECTION.LIST>      </INTERESTCOLLECTION.LIST>
                        <LEDGERCLOSINGVALUES.LIST>      </LEDGERCLOSINGVALUES.LIST>
                        <LEDGERAUDITCLASS.LIST>      </LEDGERAUDITCLASS.LIST>
                        <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                        <TDSEXEMPTIONRULES.LIST>      </TDSEXEMPTIONRULES.LIST>
                        <DEDUCTINSAMEVCHRULES.LIST>      </DEDUCTINSAMEVCHRULES.LIST>
                        <LOWERDEDUCTION.LIST>      </LOWERDEDUCTION.LIST>
                        <STXABATEMENTDETAILS.LIST>      </STXABATEMENTDETAILS.LIST>
                        <LEDMULTIADDRESSLIST.LIST>      </LEDMULTIADDRESSLIST.LIST>
                        <STXTAXDETAILS.LIST>      </STXTAXDETAILS.LIST>
                        <CHEQUERANGE.LIST>      </CHEQUERANGE.LIST>
                        <DEFAULTVCHCHEQUEDETAILS.LIST>      </DEFAULTVCHCHEQUEDETAILS.LIST>
                        <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                        <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                        <BRSIMPORTEDINFO.LIST>      </BRSIMPORTEDINFO.LIST>
                        <AUTOBRSCONFIGS.LIST>      </AUTOBRSCONFIGS.LIST>
                        <BANKURENTRIES.LIST>      </BANKURENTRIES.LIST>
                        <DEFAULTCHEQUEDETAILS.LIST>      </DEFAULTCHEQUEDETAILS.LIST>
                        <DEFAULTOPENINGCHEQUEDETAILS.LIST>      </DEFAULTOPENINGCHEQUEDETAILS.LIST>
                        <CANCELLEDPAYALLOCATIONS.LIST>      </CANCELLEDPAYALLOCATIONS.LIST>
                        <ECHEQUEPRINTLOCATION.LIST>      </ECHEQUEPRINTLOCATION.LIST>
                        <ECHEQUEPAYABLELOCATION.LIST>      </ECHEQUEPAYABLELOCATION.LIST>
                        <EDDPRINTLOCATION.LIST>      </EDDPRINTLOCATION.LIST>
                        <EDDPAYABLELOCATION.LIST>      </EDDPAYABLELOCATION.LIST>
                        <AVAILABLETRANSACTIONTYPES.LIST>      </AVAILABLETRANSACTIONTYPES.LIST>
                        <LEDPAYINSCONFIGS.LIST>      </LEDPAYINSCONFIGS.LIST>
                        <TYPECODEDETAILS.LIST>      </TYPECODEDETAILS.LIST>
                        <FIELDVALIDATIONDETAILS.LIST>      </FIELDVALIDATIONDETAILS.LIST>
                        <INPUTCRALLOCS.LIST>      </INPUTCRALLOCS.LIST>
                        <GSTCLASSFNIGSTRATES.LIST>      </GSTCLASSFNIGSTRATES.LIST>
                        <EXTARIFFDUTYHEADDETAILS.LIST>      </EXTARIFFDUTYHEADDETAILS.LIST>
                        <VOUCHERTYPEPRODUCTCODES.LIST>      </VOUCHERTYPEPRODUCTCODES.LIST>
                        </LEDGER>
                    </TALLYMESSAGE>
                    '''

            xml_final=xml_begin+xml+xml_end
            request.session['xml_final'] = xml_final
            csv_string = data.to_csv(index=False)
            request.session['data_csv'] = csv_string
            request.session['file_name'] = 'Master_Duties'
            # download_url = reverse('view:download_xml') + f'?xml_final={xml_final}&csv_string={csv_string}'

            
        # You can add further processing of the uploaded file here

        return render(request, 'GSTapp/master_ledger.html', {'message':True,'file_name':uploaded_file})

def download_xml(request):
    xml_final = request.session.get('xml_final')
    csv_string = request.session.get('data_csv')
    file_name = request.session.get('file_name')

    if not xml_final or csv_string is None:
        return HttpResponse("No files to download.", content_type='text/plain')
    
    df = pd.read_csv(io.StringIO(csv_string))

    # Create a zip file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr(file_name+".xml", xml_final)
        zf.writestr("data.csv", csv_string.encode('utf-8'))  # Ensure encoding is set to 'utf-8'
    
    zip_buffer.seek(0)

    # Serve the zip file as a download
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="files.zip"'
    return response

def download_xml_file(request):
    xml_final = request.session.get('xml_final')
    file_name = request.session.get('file_name')

    if not xml_final:
        return HttpResponse("No files to download.", content_type='text/plain')

    # Serve the zip file as a download
    response = HttpResponse(xml_final, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="{file_name+".xml" or "default.xml"}"'
    return response

def download_xml_excel(request):
    csv_string = request.session.get('data_csv')

    if csv_string is None:
        return HttpResponse("No files to download.", content_type='text/plain')
    
    df = pd.read_csv(io.StringIO(csv_string))
    excel_buffer = io.BytesIO()

    # Write the DataFrame to the Excel buffer
    with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Template', index=False)
    
    # Ensure the buffer's position is at the beginning
    excel_buffer.seek(0)

    # Serve the Excel file as a download
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
    return response

class Pay_Con_Rec(View):
    def get(self,request):
        return render(request,'GSTapp/pay_con_rec.html')
    
    def post(self,request):
        uploaded_file = request.FILES.get('upload_file_pay_con_rec')

        if uploaded_file:
            data=pd.read_excel(uploaded_file,sheet_name='TEMPLATE')

            data['DATE'] = pd.NaT

            try:
                # If 'Datetime' column is already datetime type, format it
                data['DATE'] = pd.to_datetime(data['DATE_TIME']).dt.strftime('%Y%m%d')
            except:
                # If 'Datetime' column is string type, parse and format it
                for x in range(len(data['DATE_TIME'])):
                    if '/' in data['DATE_TIME'].iloc[x]:
                        string = data['DATE_TIME'].iloc[x].split('/')
                    elif '-' in data['DATE_TIME'].iloc[x]:
                        string = data['DATE_TIME'].iloc[x].split('-')
                    else:
                        continue
                    data['DATE'].iloc[x] = string[-1] + string[-2] + string[-3]

            # Set default date for missing values
            default_date = pd.to_datetime('1998-01-01')  # Define your default date
            data['DATE'] = data['DATE'].fillna(default_date)

            for x in data:
                if ('int' in str(data[x].dtypes) or 'float' in str(data[x].dtypes))and 'Supplier_Invoice' not in x:
                    data[x]=data[x].astype("float").round(2)

            data=data.fillna('')

            xml_begin= '''
                <ENVELOPE>
                <HEADER>
                <TALLYREQUEST>Import Data</TALLYREQUEST>
                </HEADER>
                <BODY>
                <IMPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>All Masters</REPORTNAME>
                    <STATICVARIABLES>
                    <SVCURRENTCOMPANY>Demo Company</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
                <REQUESTDATA>
                '''
            xml_end='''
                </REQUESTDATA>
                </IMPORTDATA>
                </BODY>
                </ENVELOPE>
                '''
            xml=''

            for row in range(len(data)):
                xml=xml+'''
                    <TALLYMESSAGE xmlns:UDF="TallyUDF">
                         <VOUCHER REMOTEID="" VCHKEY="" VCHTYPE="'''+str(data.iloc[row]['Vch_Type'])+''' " ACTION="Create" OBJVIEW="Accounting Voucher View">
                          <DATE>'''+str(data.iloc[row]['DATE'])+'''</DATE>
                          <GUID></GUID>
                          <NARRATION>'''+str(data.iloc[row]['Narration'])+'''</NARRATION>
                          <VOUCHERTYPENAME>'''+str(data.iloc[row]['Vch_Type'])+''' </VOUCHERTYPENAME>
                          <VOUCHERNUMBER>1</VOUCHERNUMBER>
                          <PARTYLEDGERNAME>'''+str(data.iloc[row]['PartyLedgerName'])+'''</PARTYLEDGERNAME>
                          <CSTFORMISSUETYPE/>
                          <CSTFORMRECVTYPE/>
                          <FBTPAYMENTTYPE>Default</FBTPAYMENTTYPE>
                          <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>
                          <VCHGSTCLASS/>
                          <DIFFACTUALQTY>No</DIFFACTUALQTY>
                          <ISMSTFROMSYNC>No</ISMSTFROMSYNC>
                          <ASORIGINAL>No</ASORIGINAL>
                          <AUDITED>No</AUDITED>
                          <FORJOBCOSTING>No</FORJOBCOSTING>
                          <ISOPTIONAL>No</ISOPTIONAL>
                          <EFFECTIVEDATE>'''+str(data.iloc[row]['DATE'])+'''</EFFECTIVEDATE>
                          <USEFOREXCISE>No</USEFOREXCISE>
                          <ISFORJOBWORKIN>No</ISFORJOBWORKIN>
                          <ALLOWCONSUMPTION>No</ALLOWCONSUMPTION>
                          <USEFORINTEREST>No</USEFORINTEREST>
                          <USEFORGAINLOSS>No</USEFORGAINLOSS>
                          <USEFORGODOWNTRANSFER>No</USEFORGODOWNTRANSFER>
                          <USEFORCOMPOUND>No</USEFORCOMPOUND>
                          <USEFORSERVICETAX>No</USEFORSERVICETAX>
                          <ISEXCISEVOUCHER>No</ISEXCISEVOUCHER>
                          <EXCISETAXOVERRIDE>No</EXCISETAXOVERRIDE>
                          <USEFORTAXUNITTRANSFER>No</USEFORTAXUNITTRANSFER>
                          <EXCISEOPENING>No</EXCISEOPENING>
                          <USEFORFINALPRODUCTION>No</USEFORFINALPRODUCTION>
                          <ISTDSOVERRIDDEN>No</ISTDSOVERRIDDEN>
                          <ISTCSOVERRIDDEN>No</ISTCSOVERRIDDEN>
                          <ISTDSTCSCASHVCH>No</ISTDSTCSCASHVCH>
                          <INCLUDEADVPYMTVCH>No</INCLUDEADVPYMTVCH>
                          <ISSUBWORKSCONTRACT>No</ISSUBWORKSCONTRACT>
                          <ISVATOVERRIDDEN>No</ISVATOVERRIDDEN>
                          <IGNOREORIGVCHDATE>No</IGNOREORIGVCHDATE>
                          <ISSERVICETAXOVERRIDDEN>No</ISSERVICETAXOVERRIDDEN>
                          <ISISDVOUCHER>No</ISISDVOUCHER>
                          <ISEXCISEOVERRIDDEN>No</ISEXCISEOVERRIDDEN>
                          <ISEXCISESUPPLYVCH>No</ISEXCISESUPPLYVCH>
                          <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>
                          <GSTNOTEXPORTED>No</GSTNOTEXPORTED>
                          <ISVATPRINCIPALACCOUNT>No</ISVATPRINCIPALACCOUNT>
                          <ISBOENOTAPPLICABLE>No</ISBOENOTAPPLICABLE>
                          <ISSHIPPINGWITHINSTATE>No</ISSHIPPINGWITHINSTATE>
                          <ISOVERSEASTOURISTTRANS>No</ISOVERSEASTOURISTTRANS>
                          <ISCANCELLED>No</ISCANCELLED>
                          <HASCASHFLOW>No</HASCASHFLOW>
                          <ISPOSTDATED>No</ISPOSTDATED>
                          <USETRACKINGNUMBER>No</USETRACKINGNUMBER>
                          <ISINVOICE>No</ISINVOICE>
                          <MFGJOURNAL>No</MFGJOURNAL>
                          <HASDISCOUNTS>No</HASDISCOUNTS>
                          <ASPAYSLIP>No</ASPAYSLIP>
                          <ISCOSTCENTRE>No</ISCOSTCENTRE>
                          <ISSTXNONREALIZEDVCH>No</ISSTXNONREALIZEDVCH>
                          <ISEXCISEMANUFACTURERON>No</ISEXCISEMANUFACTURERON>
                          <ISBLANKCHEQUE>No</ISBLANKCHEQUE>
                          <ISVOID>No</ISVOID>
                          <ISONHOLD>No</ISONHOLD>
                          <ORDERLINESTATUS>No</ORDERLINESTATUS>
                          <VATISAGNSTCANCSALES>No</VATISAGNSTCANCSALES>
                          <VATISPURCEXEMPTED>No</VATISPURCEXEMPTED>
                          <ISVATRESTAXINVOICE>No</ISVATRESTAXINVOICE>
                          <VATISASSESABLECALCVCH>No</VATISASSESABLECALCVCH>
                          <ISVATDUTYPAID>Yes</ISVATDUTYPAID>
                          <ISDELIVERYSAMEASCONSIGNEE>No</ISDELIVERYSAMEASCONSIGNEE>
                          <ISDISPATCHSAMEASCONSIGNOR>No</ISDISPATCHSAMEASCONSIGNOR>
                          <ISDELETED>No</ISDELETED>
                          <CHANGEVCHMODE>No</CHANGEVCHMODE>
                          <ALTERID> 1</ALTERID>
                          <MASTERID> 1</MASTERID>
                          <VOUCHERKEY>184322816475144</VOUCHERKEY>
                          <EXCLUDEDTAXATIONS.LIST>      </EXCLUDEDTAXATIONS.LIST>
                          <OLDAUDITENTRIES.LIST>      </OLDAUDITENTRIES.LIST>
                          <ACCOUNTAUDITENTRIES.LIST>      </ACCOUNTAUDITENTRIES.LIST>
                          <AUDITENTRIES.LIST>      </AUDITENTRIES.LIST>
                          <DUTYHEADDETAILS.LIST>      </DUTYHEADDETAILS.LIST>
                          <SUPPLEMENTARYDUTYHEADDETAILS.LIST>      </SUPPLEMENTARYDUTYHEADDETAILS.LIST>
                          <INVOICEDELNOTES.LIST>      </INVOICEDELNOTES.LIST>
                          <INVOICEORDERLIST.LIST>      </INVOICEORDERLIST.LIST>
                          <INVOICEINDENTLIST.LIST>      </INVOICEINDENTLIST.LIST>
                          <ATTENDANCEENTRIES.LIST>      </ATTENDANCEENTRIES.LIST>
                          <ORIGINVOICEDETAILS.LIST>      </ORIGINVOICEDETAILS.LIST>
                          <INVOICEEXPORTLIST.LIST>      </INVOICEEXPORTLIST.LIST>
                          <ALLLEDGERENTRIES.LIST>
                           <OLDAUDITENTRYIDS.LIST TYPE="Number">
                            <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                           </OLDAUDITENTRYIDS.LIST>
                          <NARRATION>'''+str(data.iloc[row]['NARRATION_1'])+'''</NARRATION>
                           <LEDGERNAME>'''+str(data.iloc[row]['Dr_LedgerName'])+'''</LEDGERNAME>
                           <GSTCLASS/>
                           <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
                           <LEDGERFROMITEM>No</LEDGERFROMITEM>
                           <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                           <ISPARTYLEDGER>No</ISPARTYLEDGER>
                           <ISLASTDEEMEDPOSITIVE>Yes</ISLASTDEEMEDPOSITIVE>
                           <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                           <AMOUNT>-'''+str(data.iloc[row]['Dr_Amount'])+'''</AMOUNT>
                           <VATEXPAMOUNT>-'''+str(data.iloc[row]['Dr_Amount'])+'''</VATEXPAMOUNT>
                           <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                           <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                           <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                           <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                           <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                           <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                           <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                           <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                           <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                           <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                           <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                           <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                           <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                           <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                           <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                           <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                           <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                           <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                           <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                           <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                           <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                           <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                           <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                          </ALLLEDGERENTRIES.LIST>
                          <ALLLEDGERENTRIES.LIST>
                           <OLDAUDITENTRYIDS.LIST TYPE="Number">
                            <OLDAUDITENTRYIDS>-1</OLDAUDITENTRYIDS>
                           </OLDAUDITENTRYIDS.LIST>
                           <NARRATION>'''+str(data.iloc[row]['NARRATION_2'])+'''</NARRATION>
                           <LEDGERNAME>'''+str(data.iloc[row]['Cr_LedgerName'])+'''</LEDGERNAME>
                           <GSTCLASS/>
                           <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
                           <LEDGERFROMITEM>No</LEDGERFROMITEM>
                           <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>
                           <ISPARTYLEDGER>Yes</ISPARTYLEDGER>
                           <ISLASTDEEMEDPOSITIVE>No</ISLASTDEEMEDPOSITIVE>
                           <ISCAPVATTAXALTERED>No</ISCAPVATTAXALTERED>
                           <AMOUNT>'''+str(data.iloc[row]['Cr_Amount'])+'''</AMOUNT>
                           <VATEXPAMOUNT>'''+str(data.iloc[row]['Cr_Amount'])+'''</VATEXPAMOUNT>
                           <SERVICETAXDETAILS.LIST>       </SERVICETAXDETAILS.LIST>
                           <BANKALLOCATIONS.LIST>       </BANKALLOCATIONS.LIST>
                           <BILLALLOCATIONS.LIST>       </BILLALLOCATIONS.LIST>
                           <INTERESTCOLLECTION.LIST>       </INTERESTCOLLECTION.LIST>
                           <OLDAUDITENTRIES.LIST>       </OLDAUDITENTRIES.LIST>
                           <ACCOUNTAUDITENTRIES.LIST>       </ACCOUNTAUDITENTRIES.LIST>
                           <AUDITENTRIES.LIST>       </AUDITENTRIES.LIST>
                           <INPUTCRALLOCS.LIST>       </INPUTCRALLOCS.LIST>
                           <DUTYHEADDETAILS.LIST>       </DUTYHEADDETAILS.LIST>
                           <EXCISEDUTYHEADDETAILS.LIST>       </EXCISEDUTYHEADDETAILS.LIST>
                           <RATEDETAILS.LIST>       </RATEDETAILS.LIST>
                           <SUMMARYALLOCS.LIST>       </SUMMARYALLOCS.LIST>
                           <STPYMTDETAILS.LIST>       </STPYMTDETAILS.LIST>
                           <EXCISEPAYMENTALLOCATIONS.LIST>       </EXCISEPAYMENTALLOCATIONS.LIST>
                           <TAXBILLALLOCATIONS.LIST>       </TAXBILLALLOCATIONS.LIST>
                           <TAXOBJECTALLOCATIONS.LIST>       </TAXOBJECTALLOCATIONS.LIST>
                           <TDSEXPENSEALLOCATIONS.LIST>       </TDSEXPENSEALLOCATIONS.LIST>
                           <VATSTATUTORYDETAILS.LIST>       </VATSTATUTORYDETAILS.LIST>
                           <COSTTRACKALLOCATIONS.LIST>       </COSTTRACKALLOCATIONS.LIST>
                           <REFVOUCHERDETAILS.LIST>       </REFVOUCHERDETAILS.LIST>
                           <INVOICEWISEDETAILS.LIST>       </INVOICEWISEDETAILS.LIST>
                           <VATITCDETAILS.LIST>       </VATITCDETAILS.LIST>
                           <ADVANCETAXDETAILS.LIST>       </ADVANCETAXDETAILS.LIST>
                          </ALLLEDGERENTRIES.LIST>
                          <PAYROLLMODEOFPAYMENT.LIST>      </PAYROLLMODEOFPAYMENT.LIST>
                          <ATTDRECORDS.LIST>      </ATTDRECORDS.LIST>
                          <TEMPGSTRATEDETAILS.LIST>      </TEMPGSTRATEDETAILS.LIST>
                         </VOUCHER>
                        </TALLYMESSAGE>'''
         

            xml_final=xml_begin+xml+xml_end
            request.session['xml_final'] = xml_final
            csv_string = data.to_csv(index=False)
            request.session['data_csv'] = csv_string
            request.session['file_name'] = 'Pay_Con_Rec'
        # You can add further processing of the uploaded file here

        return render(request, 'GSTapp/pay_con_rec.html', {'message':True,'file_name':uploaded_file})

