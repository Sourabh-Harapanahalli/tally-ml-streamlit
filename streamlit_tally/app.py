"""
TALLY ML — Streamlit app.

A self-contained Excel <-> Tally-XML converter. All processing logic lives in
``tally_core`` (pure pandas/openpyxl, no web framework); this module is the
entire UI. The conversion output is byte-for-byte identical to the original
Django implementation this was ported from.

Run:  streamlit run app.py
"""

import difflib
import html
import io
import re
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import tally_core

# Must be the first Streamlit command in the script.
st.set_page_config(page_title="TALLY ML", page_icon="📒", layout="centered")

# --------------------------------------------------------------------------
# Tool catalogue — maps each converter to its upload key, conversion function
# and blank-template generator (all in tally_core).
# --------------------------------------------------------------------------
TOOLS = {
    "Purchase / Sales": {
        "upload_key": "upload_file",
        "convert": tally_core.convert_purchase_sales,
        "template": tally_core.template_purchase,
        "template_name": "Template_Purchase.xlsx",
        "help": "Purchase & Sales vouchers with GST (CGST/SGST/IGST) breakup. "
                "The template has two sheets: TEMPLATE (your data) and "
                "MASTER_LEDGER_NAME_LINK (map GST columns to your Tally ledger names).",
    },
    "Payment / Contra / Receipt": {
        "upload_key": "upload_file_pay_con_rec",
        "convert": tally_core.convert_pay_con_rec,
        "template": tally_core.template_pay_con_rec,
        "template_name": "Template_Pay_Con_Rec.xlsx",
        "help": "Payment, Contra and Receipt vouchers (Dr/Cr ledger entries).",
    },
    "Master — Ledger": {
        "upload_key": "upload_file_master_ledger",
        "convert": tally_core.convert_master_ledger,
        "template": tally_core.template_master_ledger,
        "template_name": "Template_Master_Ledger.xlsx",
        "help": "Create ledger masters (groups, GST no., opening balance, address). "
                "The REFERENCE sheet lists the valid Tally group names.",
    },
    "Master — Duties & Taxes": {
        "upload_key": "upload_file_master_duties",
        "convert": tally_core.convert_master_duties,
        "template": tally_core.template_master_duties,
        "template_name": "Template_Master_DPS.xlsx",
        "help": "Create Duties & Taxes ledgers (rate of tax, tax type).",
    },
    "Master — Purchase/Sales Ledgers": {
        "upload_key": "upload_file_master_ps",
        "convert": tally_core.convert_master_ps,
        "template": tally_core.template_master_ps,
        "template_name": "Template_Master_PS.xlsx",
        "help": "Create Purchase/Sales ledgers with nature of transaction and GST rates.",
    },
}

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@st.cache_data(show_spinner=False)
def build_template(tool_name):
    """Generate a blank template workbook and return its bytes."""
    return TOOLS[tool_name]["template"]()


# --------------------------------------------------------------------------
# Example data for each converter. These are sample rows that show the
# expected format. EXAMPLES fills the TEMPLATE sheet; EXAMPLE_AUX fills the
# secondary sheet (only Purchase/Sales uses the ledger-name mapping sheet).
# --------------------------------------------------------------------------
EXAMPLES = {
    "Purchase / Sales": [
        {"Supplier_Invoice": "INV-001", "Datetime": "01/04/2024", "Vch_Type": "Purchase",
         "PartyLedgerName": "ABC Traders", "Dr_LedgerName": "Purchase @ 18%",
         "Total_Amount": 11800, "GST_18": 10000, "Narration": "Goods purchased - 18% GST (intra-state)"},
        {"Supplier_Invoice": "INV-002", "Datetime": "02/04/2024", "Vch_Type": "Sales",
         "PartyLedgerName": "XYZ Enterprises", "Dr_LedgerName": "Sales @ 12%",
         "Total_Amount": 5600, "GST_12": 5000, "Narration": "Goods sold - 12% GST (intra-state)"},
        {"Supplier_Invoice": "INV-003", "Datetime": "03/04/2024", "Vch_Type": "Sales",
         "PartyLedgerName": "PQR Pvt Ltd", "Dr_LedgerName": "Sales @ 18% (IGST)",
         "Total_Amount": 11800, "GST_18_IGST": 10000, "Narration": "Interstate sale - 18% IGST"},
    ],
    "Payment / Contra / Receipt": [
        {"DATE_TIME": "01/04/2024", "PartyLedgerName": "ABC Traders", "Dr_LedgerName": "ABC Traders",
         "Dr_Amount": 10000, "Cr_LedgerName": "HDFC Bank", "Cr_Amount": 10000,
         "Narration": "Payment to supplier", "Vch_Type": "Payment"},
        {"DATE_TIME": "02/04/2024", "PartyLedgerName": "XYZ Enterprises", "Dr_LedgerName": "HDFC Bank",
         "Dr_Amount": 15000, "Cr_LedgerName": "XYZ Enterprises", "Cr_Amount": 15000,
         "Narration": "Receipt from customer", "Vch_Type": "Receipt"},
        {"DATE_TIME": "03/04/2024", "PartyLedgerName": "Cash", "Dr_LedgerName": "Cash",
         "Dr_Amount": 5000, "Cr_LedgerName": "HDFC Bank", "Cr_Amount": 5000,
         "Narration": "Cash withdrawn from bank", "Vch_Type": "Contra"},
    ],
    "Master — Ledger": [
        {"Ledger_Name": "ABC Traders", "Alias": "ABC", "Group_Name": "Sundry Creditors",
         "Country": "India", "State_Name": "Karnataka", "Pincode": "560001",
         "Registration_Type": "Regular", "GST_NO": "29ABCDE1234F1Z5", "Opening_Balance": 0,
         "Dr/Cr": "Cr", "Address": "123 MG Road, Bangalore"},
        {"Ledger_Name": "XYZ Enterprises", "Alias": "XYZ", "Group_Name": "Sundry Debtors",
         "Country": "India", "State_Name": "Maharashtra", "Pincode": "400001",
         "Registration_Type": "Regular", "GST_NO": "27XYZAB5678G1Z3", "Opening_Balance": 25000,
         "Dr/Cr": "Dr", "Address": "45 Marine Drive, Mumbai"},
    ],
    "Master — Duties & Taxes": [
        {"Ledger_Name": "CGST", "Group_Name": "Duties & Taxes", "Rate_of_Tax": 9, "Tax_Type": "Central Tax"},
        {"Ledger_Name": "SGST", "Group_Name": "Duties & Taxes", "Rate_of_Tax": 9, "Tax_Type": "State Tax"},
        {"Ledger_Name": "IGST", "Group_Name": "Duties & Taxes", "Rate_of_Tax": 18, "Tax_Type": "Integrated Tax"},
    ],
    "Master — Purchase/Sales Ledgers": [
        {"Ledger_Name": "Sales @ 18%", "Group_Name": "Sales Accounts",
         "Nature_of_transaction": "Sales Taxable", "RATE_OF_CGST_SGST": 18, "RATE_OF_IGST": 18},
        {"Ledger_Name": "Purchase @ 12%", "Group_Name": "Purchase Accounts",
         "Nature_of_transaction": "Purchase Taxable", "RATE_OF_CGST_SGST": 12, "RATE_OF_IGST": 12},
    ],
}

# Maps each GST column on the TEMPLATE sheet to a sample Tally ledger name on
# the MASTER_LEDGER_NAME_LINK sheet (Purchase / Sales only).
EXAMPLE_AUX = {
    "Purchase / Sales": {
        "sheet": "MASTER_LEDGER_NAME_LINK",
        "values": {
            "GST_5": "Purchase/Sales @ 5%", "GST_12": "Purchase/Sales @ 12%",
            "GST_18": "Purchase/Sales @ 18%", "GST_28": "Purchase/Sales @ 28%",
            "GST_5_IGST": "Purchase/Sales @ 5% (IGST)", "GST_12_IGST": "Purchase/Sales @ 12% (IGST)",
            "GST_18_IGST": "Purchase/Sales @ 18% (IGST)", "GST_28_IGST": "Purchase/Sales @ 28% (IGST)",
            "2.5_CGST": "CGST", "6_CGST": "CGST", "9_CGST": "CGST", "14_CGST": "CGST",
            "2.5_SGST": "SGST", "6_SGST": "SGST", "9_SGST": "SGST", "14_SGST": "SGST",
            "5_IGST": "IGST", "12_IGST": "IGST", "18_IGST": "IGST", "28_IGST": "IGST",
        },
    },
}


@st.cache_data(show_spinner=False)
def build_example(tool_name):
    """Return bytes of the blank template pre-filled with the example rows."""
    wb = load_workbook(io.BytesIO(build_template(tool_name)))

    # Fill the TEMPLATE sheet, matching each example value to its header.
    ws = wb["TEMPLATE"]
    headers = [cell.value for cell in ws[1]]
    for r, row in enumerate(EXAMPLES[tool_name], start=2):
        for header, value in row.items():
            if header in headers:
                ws.cell(row=r, column=headers.index(header) + 1, value=value)

    # Fill the secondary mapping sheet, if this tool has one.
    aux = EXAMPLE_AUX.get(tool_name)
    if aux:
        ws2 = wb[aux["sheet"]]
        for r in range(2, ws2.max_row + 1):
            key = ws2.cell(row=r, column=1).value
            if key in aux["values"]:
                ws2.cell(row=r, column=2, value=aux["values"][key])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------
# Reverse direction: Tally "All Masters" XML export  ->  Master-Ledger Excel.
#
# This mirrors the Master_Ledger view in reverse: it reads each <LEDGER> block
# from a Tally export and emits the same columns the Master — Ledger template
# uses (Ledger_Name, Alias, Group_Name, Country, State_Name, Pincode,
# Registration_Type, GST_NO, Opening_Balance, Dr/Cr, Address).
# --------------------------------------------------------------------------
MASTER_XML_TOOL = "Master XML → Excel (reverse)"

# Column order must match the Master — Ledger TEMPLATE sheet.
LEDGER_COLUMNS = [
    "Ledger_Name", "Alias", "Group_Name", "Country", "State_Name", "Pincode",
    "Registration_Type", "GST_NO", "Opening_Balance", "Dr/Cr", "Address",
]


def _tag_text(block, tag):
    """Return the unescaped, trimmed text of the first <tag>…</tag> in block."""
    m = re.search(r"<" + tag + r">(.*?)</" + tag + r">", block, re.S)
    return html.unescape(m.group(1)).strip() if m else ""


def parse_master_ledgers(xml_text):
    """Parse every <LEDGER> in a Tally All-Masters export into ledger rows."""
    rows = []
    for block in re.findall(r"<LEDGER\b.*?</LEDGER>", xml_text, re.S):
        m = re.search(r'<LEDGER\s+NAME="(.*?)"', block, re.S)
        name = html.unescape(m.group(1)).strip() if m else ""

        # Alias = 2nd <NAME> in LANGUAGENAME.LIST, when distinct from the name.
        names = [html.unescape(n).strip()
                 for n in re.findall(r"<NAME>(.*?)</NAME>", block, re.S)]
        alias = names[1] if len(names) >= 2 and names[1] and names[1] != names[0] else ""

        # Address = all <ADDRESS> lines inside ADDRESS.LIST, joined.
        addr = ""
        am = re.search(r"<ADDRESS\.LIST.*?</ADDRESS\.LIST>", block, re.S)
        if am:
            parts = [html.unescape(a).strip()
                     for a in re.findall(r"<ADDRESS>(.*?)</ADDRESS>", am.group(0), re.S)]
            addr = ", ".join(p for p in parts if p)

        state = (_tag_text(block, "LEDSTATENAME")
                 or _tag_text(block, "STATENAME")
                 or _tag_text(block, "PRIORSTATENAME"))

        # Opening balance: Tally stores Dr as negative, Cr as positive (this is
        # the inverse of the Dr/Cr -> sign mapping in the Master_Ledger view).
        ob_raw = _tag_text(block, "OPENINGBALANCE")
        opening, drcr = "", ""
        if ob_raw:
            try:
                v = float(ob_raw)
                if v != 0:
                    opening = abs(v)
                    drcr = "Dr" if v < 0 else "Cr"
            except ValueError:
                pass

        rows.append({
            "Ledger_Name": name,
            "Alias": alias,
            "Group_Name": _tag_text(block, "PARENT"),
            "Country": _tag_text(block, "COUNTRYNAME"),
            "State_Name": state,
            "Pincode": _tag_text(block, "PINCODE"),
            "Registration_Type": _tag_text(block, "GSTREGISTRATIONTYPE"),
            "GST_NO": _tag_text(block, "PARTYGSTIN"),
            "Opening_Balance": opening,
            "Dr/Cr": drcr,
            "Address": addr,
        })
    return rows


def read_tally_xml(raw_bytes):
    """Decode a Tally XML export, which is usually UTF-16, sometimes UTF-8."""
    for encoding in ("utf-16", "utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def ledgers_to_excel(rows):
    """Write ledger rows into the Master — Ledger template workbook (bytes)."""
    wb = load_workbook(io.BytesIO(build_template("Master — Ledger")))
    ws = wb["TEMPLATE"]

    # Clear any rows below the header, then write parsed data.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    headers = [cell.value for cell in ws[1]]
    for r, row in enumerate(rows, start=2):
        for header, value in row.items():
            if header in headers:
                ws.cell(row=r, column=headers.index(header) + 1, value=value)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------
# Validation: the Master — Ledger sheet must not contain duplicate GST numbers.
# --------------------------------------------------------------------------
def find_duplicate_gst(file_bytes):
    """Read the TEMPLATE sheet and locate duplicate GST_NO values.

    Returns (dataframe, duplicate_mask, duplicate_values). The mask is a
    boolean Series over the rows flagged as duplicate GST numbers; blank GST
    numbers are ignored. Returns (None, None, []) if there's no GST_NO column.
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE").fillna("")
    if "GST_NO" not in df.columns:
        return None, None, []

    # Normalise for comparison (trim + uppercase) without mutating the data.
    norm = df["GST_NO"].astype(str).str.strip().str.upper()
    non_blank = norm != ""
    dup_mask = norm.duplicated(keep=False) & non_blank
    dup_values = sorted(norm[dup_mask].unique())
    return df, dup_mask, dup_values


# Columns that are mandatory when a ledger's registration type is "Regular".
REGULAR_REQUIRED_FIELDS = ["GST_NO", "Country", "State_Name"]


def find_missing_regular_fields(df):
    """Flag rows where Registration_Type is Regular but required fields are blank.

    Returns (row_mask, detail_df) where row_mask marks the offending rows and
    detail_df lists each offending ledger with the fields it is missing.
    """
    if df is None or "Registration_Type" not in df.columns:
        return None, None

    reg = df["Registration_Type"].astype(str).str.strip().str.lower() == "regular"
    present = {f: df[f].astype(str).str.strip() != "" if f in df.columns
               else pd.Series(False, index=df.index)
               for f in REGULAR_REQUIRED_FIELDS}

    row_mask = pd.Series(False, index=df.index)
    records = []
    for idx in df.index[reg]:
        missing = [f for f in REGULAR_REQUIRED_FIELDS if not present[f][idx]]
        if missing:
            row_mask[idx] = True
            name = df.at[idx, "Ledger_Name"] if "Ledger_Name" in df.columns else ""
            records.append({
                "Row": int(idx) + 2,  # +2: 1-based + header row
                "Ledger_Name": name,
                "Missing_Fields": ", ".join(missing),
            })
    detail_df = pd.DataFrame(records) if records else None
    return row_mask, detail_df


# --------------------------------------------------------------------------
# Live data: connect to a running Tally instance over ODBC and browse it.
#
# Tally exposes an ODBC server (default port 9000) when "ODBC Server" is
# enabled. This requires (a) the Tally ODBC driver installed on the machine
# running this app and (b) Tally running with ODBC on. It therefore only works
# locally next to Tally — not on a remote/cloud host — so pyodbc is imported
# lazily and any failure degrades to setup instructions instead of crashing.
# --------------------------------------------------------------------------
ODBC_TOOL = "Tally ODBC — Live Data"

# Preset Tally collections, as Tally-dialect SQL ($field names, FROM <type>).
# The Ledgers query mirrors the Master — Ledger Excel template columns, using
# the field names verified against Tally's column metadata:
#   Ledger_Name -> $Name            Group_Name        -> $Parent
#   Country     -> $CountryName      State_Name        -> $LedStateName
#   Pincode     -> $PINCode          Registration_Type -> $GSTRegistrationType
#   GST_NO      -> $PartyGSTIN       Opening_Balance   -> $OpeningBalance
#   Address     -> $Address
# Alias and Dr/Cr have no ODBC column: aliases live in Tally's language-name
# list (not exposed via SQL), and Dr/Cr is just the sign of the opening balance.
ODBC_PRESETS = {
    "Ledgers": "SELECT $Name, $Parent, $CountryName, $LedStateName, $PINCode, "
               "$GSTRegistrationType, $PartyGSTIN, $OpeningBalance, "
               "$_ClosingBalance, $Address FROM Ledger",
    "Groups": "SELECT $Name, $Parent FROM Group",
    "Stock Items": "SELECT $Name, $Parent, $BaseUnits, $ClosingBalance, "
                   "$ClosingValue FROM StockItem",
    "Cost Centres": "SELECT $Name, $Parent FROM CostCentre",
    "Voucher Types": "SELECT $Name, $Parent, $NumberingMethod FROM VoucherType",
    "Vouchers": "SELECT $Date, $VoucherTypeName, $VoucherNumber, "
                "$PartyLedgerName, $Amount FROM Voucher",
}


def list_odbc_drivers():
    """Return the ODBC driver names registered on this machine (or raise)."""
    import pyodbc  # lazy: optional dependency
    return list(pyodbc.drivers())


def build_odbc_conn_str(mode, dsn, driver, host, port):
    """Build a pyodbc connection string from the UI inputs."""
    if mode == "DSN":
        return f"DSN={dsn}"
    return f"DRIVER={{{driver}}};SERVER={host};PORT={port}"


def run_odbc_query(conn_str, sql):
    """Run a query against Tally over ODBC and return a DataFrame.

    Raises ImportError if pyodbc is unavailable, or pyodbc.Error on connection/
    query failure — the caller surfaces these as friendly messages.
    """
    import pyodbc  # lazy: optional dependency, only needed for this section

    conn = pyodbc.connect(conn_str, autocommit=True, timeout=10)
    try:
        cursor = conn.cursor()
        cursor.execute(sql)
        columns = [d[0] for d in cursor.description] if cursor.description else []
        rows = [tuple(r) for r in cursor.fetchall()]
        return pd.DataFrame.from_records(rows, columns=columns)
    finally:
        conn.close()


def list_collection_columns(conn_str, table):
    """Return the available column names for a Tally collection (no data).

    Uses the ODBC metadata call cursor.columns(); falls back to reading the
    cursor description of a 'SELECT * FROM <table>' if metadata is empty.
    """
    import pyodbc  # lazy: optional dependency

    conn = pyodbc.connect(conn_str, autocommit=True, timeout=10)
    try:
        cursor = conn.cursor()
        cols = [row.column_name for row in cursor.columns(table=table)]
        if not cols:
            cursor.execute(f"SELECT * FROM {table}")
            cols = [d[0] for d in cursor.description] if cursor.description else []
        return cols
    finally:
        conn.close()


# --------------------------------------------------------------------------
# Ledger-name validation for Purchase / Sales.
#
# Tally imports are case-sensitive and need EXACT ledger names — a single typo
# or wrong case makes the whole XML import fail. These helpers fetch the live
# ledger names from Tally (ODBC) and compare them against the PartyLedgerName /
# Dr_LedgerName columns, suggesting the closest real name for any mismatch.
# --------------------------------------------------------------------------
PS_LEDGER_COLUMNS = ["PartyLedgerName", "Dr_LedgerName"]
# Payment / Contra / Receipt also has a credit-side ledger.
PCR_LEDGER_COLUMNS = ["PartyLedgerName", "Dr_LedgerName", "Cr_LedgerName"]


def fetch_tally_ledger_names(conn_str):
    """Return the list of ledger names from a running Tally instance (ODBC)."""
    df = run_odbc_query(conn_str, "SELECT $Name FROM Ledger")
    if df.empty:
        return []
    return [str(v).strip() for v in df.iloc[:, 0].tolist() if str(v).strip()]


def collect_ledger_names(file_bytes, columns):
    """Sorted unique non-blank ledger-name values used in the given columns."""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE").fillna("")
    names = set()
    for col in columns:
        if col in df.columns:
            for v in df[col].astype(str):
                v = v.strip()
                if v:
                    names.add(v)
    return sorted(names)


# Dr/Cr ledger columns that must never be blank, per converter.
REQUIRED_LEDGER_COLUMNS = {
    "Purchase / Sales": ["Dr_LedgerName"],
    "Payment / Contra / Receipt": ["Dr_LedgerName", "Cr_LedgerName"],
}


def find_empty_ledger_rows(file_bytes, columns):
    """Find rows where any of the given Dr/Cr ledger columns is blank.

    An empty Dr/Cr ledger name produces ``<LEDGERNAME></LEDGERNAME>`` in the
    XML, which Tally rejects. Fully-blank rows (every checked column empty) are
    ignored — they are just unused template rows, not partial data. Returns
    ``(detail_df, cols_checked)``: detail_df lists the offending rows (Row +
    each checked column) or None if none are blank.
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE").fillna("")
    cols = [c for c in columns if c in df.columns]
    if not cols:
        return None, cols

    # A blank in any required ledger column is a problem...
    any_blank = pd.Series(False, index=df.index)
    for c in cols:
        any_blank = any_blank | (df[c].astype(str).str.strip() == "")

    # ...unless the ENTIRE row is empty (an unused template row, not a voucher).
    row_all_blank = (df.astype(str).apply(lambda s: s.str.strip())
                     == "").all(axis=1)

    bad_mask = any_blank & ~row_all_blank
    if not bad_mask.any():
        return None, cols

    records = []
    for idx in df.index[bad_mask]:
        rec = {"Row": int(idx) + 2}  # +2: 1-based + header row
        for c in cols:
            rec[c] = df.at[idx, c]
        records.append(rec)
    return pd.DataFrame(records), cols


def _norm_name(s):
    """Lower-case and collapse whitespace for case-insensitive comparison."""
    return " ".join(str(s).lower().split())


def find_dates_out_of_range(file_bytes, date_from, date_to,
                            date_col="Datetime", id_col="Supplier_Invoice"):
    """Find voucher dates outside the [date_from, date_to] range.

    ``date_col`` is the date column to check (``Datetime`` for Purchase/Sales,
    ``DATE_TIME`` for Payment/Contra/Receipt) and ``id_col`` an optional
    identifier column shown in the detail table. Dates are parsed day-first
    (DD/MM/YYYY) with the same robust parser the converter uses, so string /
    datetime / other cell types all work. Returns
    ``(detail_df, unparsed_df)``: detail_df lists rows whose date is outside the
    range (or None); unparsed_df lists non-blank rows whose date could not be
    parsed at all (or None).
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE")
    if date_col not in df.columns:
        return None, None

    raw = df[date_col]
    non_blank = raw.notna() & (raw.astype(str).str.strip() != "")
    # Day-first (DD/MM/YYYY), matching the converter exactly.
    parsed = tally_core.parse_ddmmyyyy(raw)
    lo, hi = pd.Timestamp(date_from), pd.Timestamp(date_to)

    out_mask = non_blank & parsed.notna() & ((parsed < lo) | (parsed > hi))
    unparsed_mask = non_blank & parsed.isna()

    def _rows(mask):
        records = []
        for idx in df.index[mask]:
            rec = {"Row": int(idx) + 2}  # +2: 1-based + header row
            if id_col and id_col in df.columns:
                rec[id_col] = df.at[idx, id_col]
            rec[date_col] = str(df.at[idx, date_col])
            records.append(rec)
        return pd.DataFrame(records) if records else None

    return _rows(out_mask), _rows(unparsed_mask)


def find_excel_converted_dates(file_bytes, date_col, id_col=None):
    """Detect date cells Excel stored as real dates instead of literal text.

    When a file is filled using the current template the date column is Text, so
    ``01/04/2025`` stays the exact string. If cells come back as real datetime
    values it means Excel auto-converted them on entry — and for ambiguous days
    (<= 12) it may have swapped day/month (dd/mm -> mm/dd). Those values can no
    longer be trusted. Returns ``(detail_df, n_total)`` where detail_df samples
    the affected rows, or ``(None, 0)`` if the column is clean (all text/blank).
    """
    import datetime as _dt
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE")
    if date_col not in df.columns:
        return None, 0

    def _is_real_date(v):
        return isinstance(v, (pd.Timestamp, _dt.datetime, _dt.date)) and not pd.isna(v)

    mask = df[date_col].map(_is_real_date)
    if not mask.any():
        return None, 0

    records = []
    for idx in df.index[mask][:50]:  # sample: enough to show the problem
        rec = {"Row": int(idx) + 2}  # +2: 1-based + header row
        if id_col and id_col in df.columns:
            rec[id_col] = df.at[idx, id_col]
        rec[date_col] = str(df.at[idx, date_col])
        records.append(rec)
    return pd.DataFrame(records), int(mask.sum())


def find_unparsable_dates(file_bytes, date_col, id_col=None):
    """List non-blank date rows the strict parser can't read, with a suggestion.

    Each item is a dict: ``row_index`` (0-based), ``Row`` (Excel row), ``raw``
    (the original value), ``suggestion`` (a recovered pandas Timestamp or None)
    and, if available, ``id`` (the identifier column value).
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="TEMPLATE")
    if date_col not in df.columns:
        return []
    raw = df[date_col]
    non_blank = raw.notna() & (raw.astype(str).str.strip() != "")
    parsed = tally_core.parse_ddmmyyyy(raw)
    mask = non_blank & parsed.isna()

    rows = []
    for idx in df.index[mask]:
        rec = {
            "row_index": int(idx),
            "Row": int(idx) + 2,  # +2: 1-based + header row
            "raw": str(df.at[idx, date_col]),
            "suggestion": tally_core.suggest_date(df.at[idx, date_col]),
        }
        if id_col and id_col in df.columns:
            rec["id"] = df.at[idx, id_col]
        rows.append(rec)
    return rows


def apply_date_corrections(file_bytes, date_col, corrections):
    """Rewrite date cells (keyed by 0-based row index) to 'dd/mm/yyyy' text.

    Only the date column on the TEMPLATE sheet is touched; the cell is stored as
    Text so the corrected value can't be re-mangled by Excel.
    """
    if not corrections:
        return file_bytes
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb["TEMPLATE"]
    headers = [c.value for c in ws[1]]
    if date_col not in headers:
        return file_bytes
    ci = headers.index(date_col) + 1
    for row_index, value in corrections.items():
        cell = ws.cell(row=int(row_index) + 2, column=ci)
        cell.value = value
        cell.number_format = "@"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def date_resolver_ui(file_bytes, date_col, id_col, key_prefix):
    """Confirm/repair unreadable dates before conversion (blocks until clean).

    Suggests a compatible DD/MM/YYYY value for each unreadable date and offers a
    calendar picker to override it. Returns ``(convert_bytes, blocked)``;
    ``blocked`` stays True until every unreadable date has been resolved.
    """
    rows = find_unparsable_dates(file_bytes, date_col, id_col)
    if not rows:
        return file_bytes, False

    corr_key = f"{key_prefix}_datecorr"
    corrections = st.session_state.setdefault(corr_key, {})
    n_sugg = sum(1 for r in rows if r["suggestion"] is not None)

    st.subheader("Step 2b — Fix unreadable dates")
    st.warning(
        f"⚠️ **{len(rows)}** date(s) can't be read as DD/MM/YYYY. Confirm the "
        f"suggested fix or pick a date from the calendar. **Conversion is "
        "paused until every date is resolved.**"
    )

    # Fast path: accept every suggestion at once.
    if n_sugg and st.button(f"✨ Apply all {n_sugg} suggested date(s)",
                            key=f"{key_prefix}_date_all", type="primary"):
        for r in rows:
            if r["suggestion"] is not None:
                corrections[r["row_index"]] = r["suggestion"].strftime("%d/%m/%Y")
        st.rerun()

    # Manual review / override, one calendar per row.
    with st.form(f"{key_prefix}_dateform"):
        st.caption("Original value → confirm or change the date, then Apply.")
        picks = {}
        for i, r in enumerate(rows):
            c0, c1, c2 = st.columns([1, 3, 2])
            c0.markdown(f"Row **{r['Row']}**")
            ident = f"**{r['id']}**  " if r.get("id") else ""
            c1.markdown(f"{ident}`{r['raw']}`")
            staged = corrections.get(r["row_index"])
            if staged:
                default = pd.to_datetime(staged, dayfirst=True).date()
            elif r["suggestion"] is not None:
                default = r["suggestion"].date()
            else:
                default = None
            picks[r["row_index"]] = c2.date_input(
                "Date", value=default, format="DD/MM/YYYY",
                min_value=date(2000, 1, 1), max_value=date(2100, 12, 31),
                key=f"{key_prefix}_dp_{i}", label_visibility="collapsed")
        if st.form_submit_button("✅ Apply these dates"):
            for ri, d in picks.items():
                if d is not None:
                    corrections[ri] = d.strftime("%d/%m/%Y")
            st.rerun()

    convert_bytes = apply_date_corrections(file_bytes, date_col, corrections)
    remaining = find_unparsable_dates(convert_bytes, date_col, id_col)
    if remaining:
        st.info(f"⏸️ {len(remaining)} date(s) still unresolved — apply a "
                "suggestion or pick a date above to continue.")
        return convert_bytes, True
    st.success("✅ All dates are now readable.")
    return convert_bytes, False


def closest_ledger_names(name, tally_names, n=6):
    """Return up to n closest Tally ledger names to `name`, best first.

    Case-INSENSITIVE (Tally names are often title-case while uploads are
    upper-case) and word-aware: the score blends a character-level similarity
    with a token-overlap (Jaccard) score, so reordered or extra words (e.g. an
    extra "AND") still rank the right ledger highly. Always returns the best
    candidates so the user has options to pick from.
    """
    q = _norm_name(name)
    q_tokens = set(q.split())
    scored = []
    for cand in tally_names:
        c = _norm_name(cand)
        seq = difflib.SequenceMatcher(None, q, c).ratio()
        c_tokens = set(c.split())
        union = q_tokens | c_tokens
        jacc = len(q_tokens & c_tokens) / len(union) if union else 0.0
        score = 0.6 * seq + 0.4 * jacc
        scored.append((score, cand))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [cand for _score, cand in scored[:n]]


def apply_ledger_corrections(file_bytes, corrections, columns=PS_LEDGER_COLUMNS):
    """Return new .xlsx bytes with the given ledger columns' cells renamed.

    Only the named ledger columns on the TEMPLATE sheet are touched; everything
    else (including any MASTER_LEDGER_NAME_LINK sheet) is preserved verbatim.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb["TEMPLATE"]
    headers = [c.value for c in ws[1]]
    for col in columns:
        if col in headers:
            ci = headers.index(col) + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                key = str(cell.value).strip() if cell.value is not None else ""
                if key in corrections:
                    cell.value = corrections[key]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def odbc_connection_form(prefix):
    """Render Tally ODBC connection inputs and return a pyodbc connection string."""
    c1, c2 = st.columns(2)
    host = c1.text_input("Host", value="localhost", key=f"{prefix}_host")
    port = c2.number_input("Port", value=9000, step=1, key=f"{prefix}_port")
    mode = st.radio("Connection style", ["DSN", "Driver name"],
                    horizontal=True, key=f"{prefix}_mode")
    if mode == "DSN":
        dsn = st.text_input("ODBC DSN name", value="TallyODBC64_9000", key=f"{prefix}_dsn")
        return build_odbc_conn_str("DSN", dsn, "", host, int(port))
    driver = st.text_input("Driver name (exact)", value="Tally ODBC Driver64",
                           key=f"{prefix}_driver")
    return build_odbc_conn_str("Driver", "", driver, host, int(port))


def date_range_ui(key_prefix):
    """Optional voucher-date range gate (before upload).

    ``key_prefix`` namespaces the Streamlit widget keys so each converter keeps
    its own state. Returns ``(enabled, date_from, date_to)``; when enabled,
    every voucher's date must fall within the range or conversion is blocked.
    """
    st.subheader("Step 1b — Voucher date range (optional)")
    enabled = st.checkbox(
        "Restrict voucher dates to a range (your Tally data period)",
        key=f"{key_prefix}_date_enable",
        help="Tally rejects vouchers dated outside the company's open period. "
             "Turn this on to catch out-of-range dates before converting.",
    )
    if not enabled:
        return False, None, None

    c1, c2 = st.columns(2)
    date_from = c1.date_input("From date", key=f"{key_prefix}_date_from",
                              format="DD/MM/YYYY")
    date_to = c2.date_input("To date", key=f"{key_prefix}_date_to",
                            format="DD/MM/YYYY")
    if date_from and date_to and date_from > date_to:
        st.warning("‘From date’ is after ‘To date’ — please fix the range.")
    return True, date_from, date_to


def ledger_check_ui(file_bytes, columns, key_prefix, columns_label):
    """Verify a converter's ledger names against Tally (ODBC), before conversion.

    ``columns`` are the TEMPLATE columns to validate, ``key_prefix`` namespaces
    the Streamlit widget/session keys so multiple converters don't collide, and
    ``columns_label`` is the human-readable list shown in the expander header.

    Returns ``(convert_bytes, blocked)``:

    * ``convert_bytes`` — the (possibly corrected) workbook bytes to convert.
    * ``blocked`` — True while conversion must wait: the Tally check hasn't
      been run yet (and not skipped), or there are unresolved mismatches the
      user hasn't chosen to ignore. The caller must not convert when True.
    """
    convert_bytes = file_bytes
    ledgers_key = f"{key_prefix}_tally_ledgers"
    corr_key = f"{key_prefix}_corrections"
    st.subheader("Step 2a — Verify ledger names against Tally")
    with st.expander(f"🔗 Check {columns_label} against your Tally company",
                     expanded=True):
        st.caption(
            "Tally imports are **case-sensitive** and need the exact ledger "
            "names — a small typo or wrong case makes the whole import fail. "
            "Connect to a running Tally (ODBC, same machine) to catch problems "
            "before converting."
        )
        conn_str = odbc_connection_form(f"{key_prefix}_chk")
        st.caption(f"Connection string: `{conn_str}`")

        if st.button("🔍 Fetch ledger names from Tally & check",
                     key=f"{key_prefix}_chk_fetch"):
            try:
                names = fetch_tally_ledger_names(conn_str)
                st.session_state[ledgers_key] = names
                st.session_state[corr_key] = {}
                if not names:
                    st.warning("Connected, but Tally returned no ledgers.")
            except ImportError:
                st.error("`pyodbc` is not installed. Run `pip install pyodbc`.")
            except Exception as exc:  # noqa: BLE001
                st.session_state.pop(ledgers_key, None)
                st.error(f"Could not fetch ledgers from Tally: {exc}")
                st.caption(
                    "Check that Tally is open with ODBC enabled and the port "
                    "matches, and that the driver name / DSN and bitness are right."
                )

        tally_names = st.session_state.get(ledgers_key)
        if not tally_names:
            # Conversion waits until the check is run — unless explicitly skipped.
            skip = st.checkbox(
                "Skip Tally check (convert without validating ledger names)",
                key=f"{key_prefix}_chk_skip")
            if skip:
                st.caption("⚠️ Ledger names will NOT be validated against Tally.")
                return convert_bytes, False
            st.info("⏸️ Conversion is paused — fetch ledger names from Tally above "
                    "to validate, or tick “Skip Tally check”.")
            return convert_bytes, True

        corrections = st.session_state.setdefault(corr_key, {})
        tally_set = set(tally_names)
        used = collect_ledger_names(file_bytes, columns)

        unresolved = []
        for name in used:
            effective = corrections.get(name, name)
            if effective in tally_set:
                continue
            unresolved.append((name, effective, closest_ledger_names(effective, tally_names)))

        st.caption(f"Loaded **{len(tally_names)}** ledger(s) from Tally · "
                   f"checking **{len(used)}** name(s) used in your file.")

        if corrections:
            st.caption(f"✏️ {len(corrections)} fix(es) staged — they'll be applied "
                       "to the file before conversion.")

        if not unresolved:
            st.success("✅ All ledger names match Tally — converting below.")
            corrected = (apply_ledger_corrections(file_bytes, corrections, columns)
                         if corrections else convert_bytes)
            return corrected, False

        # Offer a one-click "apply all suggestions" for names that have one.
        fixable = [u for u in unresolved if u[2]]
        if fixable and st.button(f"✨ Apply all {len(fixable)} suggested fix(es)",
                                 key=f"{key_prefix}_chk_fix_all", type="primary"):
            for name, _eff, sugg in fixable:
                corrections[name] = sugg[0]
            st.rerun()

        st.warning(f"⚠️ **{len(unresolved)}** ledger name(s) don't match Tally. "
                   f"Pick the right ledger (search across all {len(tally_names)} "
                   "Tally ledgers) and click Fix.")

        sorted_names = sorted(tally_names, key=str.lower)
        for i, (name, effective, suggestions) in enumerate(unresolved):
            shown = name if effective == name else f"{name}  →  {effective}"
            st.markdown(f"❌ `{shown}` — pick the correct ledger:")

            # Suggestions first (best match pre-selected), then every other
            # ledger, all in their original Tally case.
            seen = set(suggestions)
            options = suggestions + [n for n in sorted_names if n not in seen]

            sc1, sc2, sc3 = st.columns([3, 4, 1])
            # Case-INSENSITIVE substring search; the list keeps original case
            # and narrows as you type (press Enter to apply the filter).
            query = sc1.text_input(
                "Search", key=f"{key_prefix}_chk_search_{i}",
                placeholder="🔎 search (case-insensitive)…",
                label_visibility="collapsed")
            if query:
                ql = query.lower()
                filtered = [n for n in options if ql in n.lower()]
            else:
                filtered = options
            if not filtered:
                sc1.caption("No ledger contains that text.")
                filtered = options

            choice = sc2.selectbox(
                "Closest Tally ledger", filtered,
                key=f"{key_prefix}_chk_sugg_{i}", label_visibility="collapsed")
            if sc3.button("Fix", key=f"{key_prefix}_chk_fix_{i}"):
                corrections[name] = choice
                st.rerun()

        # Apply whatever has been staged so far so partial fixes still take effect.
        convert_bytes = (apply_ledger_corrections(file_bytes, corrections, columns)
                         if corrections else file_bytes)

        # Conversion stays paused while names are unresolved, unless the user
        # explicitly opts to convert anyway (e.g. ledgers created later in Tally).
        convert_anyway = st.checkbox(
            "Convert anyway, ignoring the unmatched names above",
            key=f"{key_prefix}_chk_convert_anyway")
        if not convert_anyway:
            st.info("⏸️ Conversion is paused until every name is fixed "
                    "(or tick “Convert anyway”).")

    return convert_bytes, (not convert_anyway)


def ps_ledger_check_ui(file_bytes):
    """Ledger-name check for Purchase / Sales (PartyLedgerName / Dr_LedgerName)."""
    return ledger_check_ui(
        file_bytes, PS_LEDGER_COLUMNS, "ps",
        "PartyLedgerName / Dr_LedgerName")


def pcr_ledger_check_ui(file_bytes):
    """Ledger-name check for Payment / Contra / Receipt (Party / Dr / Cr)."""
    return ledger_check_ui(
        file_bytes, PCR_LEDGER_COLUMNS, "pcr",
        "PartyLedgerName / Dr_LedgerName / Cr_LedgerName")


# --------------------------------------------------------------------------
# UI
# --------------------------------------------------------------------------
st.sidebar.title("📒 TALLY ML")
st.sidebar.caption("Excel → Tally XML converter")
tool_name = st.sidebar.radio(
    "Choose a converter", list(TOOLS.keys()) + [MASTER_XML_TOOL, ODBC_TOOL])

# ==========================================================================
# Live tool: browse a running Tally instance over ODBC.
# ==========================================================================
if tool_name == ODBC_TOOL:
    st.title(ODBC_TOOL)
    st.info(
        "Connect to a **running Tally** instance over its ODBC server and view "
        "the data live in this dashboard. This works only when the app runs on "
        "the same machine as Tally (with the Tally ODBC driver installed)."
    )

    with st.expander("ℹ️ How to enable ODBC in Tally", expanded=False):
        st.markdown(
            """
1. Open **Tally** (TallyPrime / Tally.ERP 9) and load your company.
2. Go to **F1: Help → Settings → Connectivity** (or **F12: Configure → Advanced**).
3. Set **ODBC Server** to **Yes** and note the **port** (default **9000**).
4. Keep Tally open while you use this section.
5. `pip install pyodbc` if it isn't already available.

If you see **“HY000 — the driver did not supply an error”**, it's usually a
wrong **driver name** or a **32-bit/64-bit mismatch** — use *List installed
ODBC drivers* below to find the exact name.
"""
        )

    st.subheader("Step 1 — Connect")
    col_a, col_b = st.columns(2)
    host = col_a.text_input("Host", value="localhost", key="tally_host")
    port = col_b.number_input("Port", value=9000, step=1, key="tally_port")

    if st.button("🔍 List installed ODBC drivers"):
        try:
            drivers = list_odbc_drivers()
            if drivers:
                st.write("Installed ODBC drivers on this machine:")
                st.code("\n".join(drivers))
                tally_like = [d for d in drivers if "tally" in d.lower()]
                if tally_like:
                    st.success("Tally driver(s) found: " +
                               ", ".join(f"`{d}`" for d in tally_like) +
                               " — copy the exact name into the field below.")
                else:
                    st.warning(
                        "No Tally driver found. Install the Tally ODBC driver "
                        "matching your Python bitness."
                    )
            else:
                st.warning("No ODBC drivers are registered on this machine.")
        except ImportError:
            st.error("`pyodbc` is not installed. Run `pip install pyodbc`.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not list drivers: {exc}")

    odbc_mode = st.radio("ODBC connection style", ["DSN", "Driver name"],
                         horizontal=True, key="odbc_mode")
    if odbc_mode == "DSN":
        dsn = st.text_input("ODBC DSN name", value="TallyODBC64_9000",
                            key="odbc_dsn")
        conn_str = build_odbc_conn_str("DSN", dsn, "", host, int(port))
    else:
        driver = st.text_input("Driver name (exact)",
                               value="Tally ODBC Driver64", key="odbc_driver")
        conn_str = build_odbc_conn_str("Driver", "", driver, host, int(port))
    st.caption(f"Connection string: `{conn_str}`")

    if st.button("🔌 Test connection"):
        try:
            df = run_odbc_query(conn_str, "SELECT $Name FROM Company")
            names = [str(v) for v in df.iloc[:, 0].tolist()] if not df.empty else []
            st.success(f"Connected. Open compan(y/ies): {', '.join(names) or '—'}")
        except ImportError:
            st.error("`pyodbc` is not installed. Run `pip install pyodbc`.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not connect: {exc}")
            st.caption(
                "Check that Tally is open with ODBC enabled and the port "
                "matches. Confirm the driver name / DSN and bitness."
            )

    st.divider()

    # ---- Browse data ----------------------------------------------------
    st.subheader("Step 2 — Browse data")
    preset = st.selectbox("Pick a collection", list(ODBC_PRESETS.keys()))

    # Table name = the word after FROM in the preset SQL (Ledger, Group, …).
    table = ODBC_PRESETS[preset].split("FROM")[-1].strip().split()[0]
    if st.button(f"🧾 Show available columns for {table}"):
        try:
            cols = list_collection_columns(conn_str, table)
            if cols:
                st.write(f"{len(cols)} column(s) available on `{table}`:")
                st.code(", ".join(cols))
            else:
                st.warning("The driver returned no column metadata for this table.")
        except ImportError:
            st.error("`pyodbc` is not installed. Run `pip install pyodbc`.")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not list columns: {exc}")

    sql = st.text_area("SQL query (Tally dialect)", value=ODBC_PRESETS[preset],
                       height=100, key=f"odbc_sql_{preset}")

    if st.button("▶️ Run query", type="primary"):
        with st.spinner("Querying Tally…"):
            try:
                df = run_odbc_query(conn_str, sql)
                if df.empty:
                    st.warning("Query ran but returned no rows.")
                else:
                    st.success(f"{len(df)} row(s).")
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    out = io.BytesIO()
                    df.to_excel(out, index=False)
                    slug = preset.lower().replace(" ", "_")
                    c1, c2 = st.columns(2)
                    c1.download_button(
                        "⬇️ Excel (.xlsx)", data=out.getvalue(),
                        file_name=f"tally_{slug}.xlsx", mime=XLSX_MIME,
                    )
                    c2.download_button(
                        "⬇️ CSV", data=df.to_csv(index=False).encode("utf-8"),
                        file_name=f"tally_{slug}.csv", mime="text/csv",
                    )
            except ImportError:
                st.error("`pyodbc` is not installed. Run `pip install pyodbc`.")
            except Exception as exc:  # noqa: BLE001
                st.error(f"Query failed: {exc}")

    st.stop()

# ==========================================================================
# Reverse tool: Tally Master XML  ->  Master-Ledger Excel.
# ==========================================================================
if tool_name == MASTER_XML_TOOL:
    st.title(MASTER_XML_TOOL)
    st.info(
        "Upload a Tally **All Masters** XML export (e.g. `Master.xml`). Each "
        "ledger is extracted into the **Master — Ledger** Excel format "
        "(Ledger_Name, Alias, Group_Name, Country, State_Name, Pincode, "
        "Registration_Type, GST_NO, Opening_Balance, Dr/Cr, Address)."
    )

    st.subheader("Step 1 — Export your masters from Tally as XML")
    st.markdown(
        """
1. In Tally, navigate to **Display**.
2. Open **List of Accounts**.
3. Press **Alt + E** to **Export**.
4. Change the **Format** to **XML**.
5. Check the **output file location** (note the folder where the file is saved).
6. Click **Export**.
7. **Upload the XML file** below.
"""
    )

    st.divider()

    st.subheader("Step 2 — Upload your Tally Master XML")
    xml_file = st.file_uploader(
        "Upload the Tally All-Masters XML export",
        type=["xml"],
        key="uploader_master_xml",
    )

    if xml_file is not None:
        with st.spinner("Parsing ledgers…"):
            try:
                xml_text = read_tally_xml(xml_file.getvalue())
                rows = parse_master_ledgers(xml_text)

                if not rows:
                    st.error(
                        "No <LEDGER> records were found. Make sure this is a "
                        "Tally 'All Masters' XML export."
                    )
                else:
                    st.success(f"Done! Found {len(rows)} ledger(s).")
                    df = pd.DataFrame(rows, columns=LEDGER_COLUMNS)
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    st.download_button(
                        "⬇️ Download Excel (Master — Ledger format)",
                        data=ledgers_to_excel(rows),
                        file_name="Master_Ledger_from_XML.xlsx",
                        mime=XLSX_MIME,
                    )
            except Exception as exc:  # noqa: BLE001
                st.error(f"Conversion failed: {exc}")
                st.exception(exc)

    st.stop()

tool = TOOLS[tool_name]

st.title(tool_name)
st.info(tool["help"])

# ---- Step 1: download the blank template -------------------------------
st.subheader("Step 1 — Download the blank template")
try:
    template_bytes = build_template(tool_name)
    st.download_button(
        label=f"⬇️ Download {tool['template_name']}",
        data=template_bytes,
        file_name=tool["template_name"],
        mime=XLSX_MIME,
    )
except Exception as exc:  # noqa: BLE001
    st.error(f"Could not generate template: {exc}")

# ---- Example: show & download a pre-filled template --------------------
with st.expander("👀 See an example (filled template)"):
    st.caption(
        "Sample rows showing the expected format. Use these as a guide, then "
        "replace them with your own data in the blank template above."
    )
    example_rows = EXAMPLES.get(tool_name)
    if example_rows:
        st.dataframe(pd.DataFrame(example_rows), use_container_width=True, hide_index=True)
        try:
            st.download_button(
                label=f"⬇️ Download filled example — {tool['template_name']}",
                data=build_example(tool_name),
                file_name=f"EXAMPLE_{tool['template_name']}",
                mime=XLSX_MIME,
                key=f"example_{tool['upload_key']}",
            )
        except Exception as exc:  # noqa: BLE001
            st.warning(f"Could not generate the example file: {exc}")

# ---- Optional voucher date range, before upload ------------------------
date_enabled, date_from, date_to = (False, None, None)
if tool_name == "Purchase / Sales":
    date_enabled, date_from, date_to = date_range_ui("ps")
elif tool_name == "Payment / Contra / Receipt":
    date_enabled, date_from, date_to = date_range_ui("pcr")

st.divider()

# ---- Step 2: upload the filled template and convert --------------------
st.subheader("Step 2 — Upload your filled template")
uploaded = st.file_uploader(
    "Upload the completed Excel file",
    type=["xlsx", "xls"],
    key=f"uploader_{tool['upload_key']}",
)

if uploaded is not None:
    block_conversion = False

    # ---- Validation (Master — Ledger) ----------------------------------
    if tool_name == "Master — Ledger":
        try:
            val_df, dup_mask, dup_values = find_duplicate_gst(uploaded.getvalue())
        except Exception as exc:  # noqa: BLE001
            val_df, dup_mask, dup_values = None, None, []
            st.warning(f"Could not run GST validation: {exc}")

        # (a) Duplicate GST numbers — warning only (does not block).
        if dup_values:
            st.warning(
                f"⚠️ Found **{len(dup_values)}** duplicate GST number(s) across "
                f"**{int(dup_mask.sum())}** row(s). Each GST number should be "
                "unique. The duplicate rows are highlighted below — review them "
                "before converting."
            )
            st.markdown("**Duplicate GST numbers:** " +
                        ", ".join(f"`{g}`" for g in dup_values))

            dup_rows = val_df[dup_mask]
            styled = dup_rows.style.apply(
                lambda col: ["background-color: #ffd6d6"] * len(col)
                if col.name == "GST_NO" else [""] * len(col),
                axis=0,
            )
            st.dataframe(styled, use_container_width=True)
        elif val_df is not None:
            st.success("✅ No duplicate GST numbers found.")

        # (b) Regular ledgers must have GST_NO, Country and State_Name — blocks.
        reg_mask, reg_detail = find_missing_regular_fields(val_df)
        if reg_detail is not None:
            block_conversion = True
            st.error(
                f"🚫 **{len(reg_detail)}** ledger(s) have **Registration_Type = "
                "Regular** but are missing required field(s). For Regular "
                "ledgers, **GST_NO**, **Country** and **State_Name** are "
                "mandatory. Fix the rows below, then re-upload."
            )
            missing_cols = ["GST_NO", "Country", "State_Name"]
            bad_rows = val_df[reg_mask]
            styled_missing = bad_rows.style.apply(
                lambda col: [
                    "background-color: #ffd6d6" if (col.name in missing_cols
                                                    and str(v).strip() == "")
                    else "" for v in col
                ],
                axis=0,
            )
            st.dataframe(styled_missing, use_container_width=True)
            st.markdown("**Rows needing attention:**")
            st.dataframe(reg_detail, use_container_width=True, hide_index=True)

    # ---- Validation: no empty Dr/Cr ledger names (Purchase/Sales & PCR) ----
    if tool_name in REQUIRED_LEDGER_COLUMNS:
        try:
            empty_df, checked_cols = find_empty_ledger_rows(
                uploaded.getvalue(), REQUIRED_LEDGER_COLUMNS[tool_name])
        except Exception as exc:  # noqa: BLE001
            empty_df, checked_cols = None, []
            st.warning(f"Could not check for empty ledger names: {exc}")

        if empty_df is not None:
            block_conversion = True
            cols_txt = " / ".join(checked_cols)
            st.error(
                f"🚫 **{len(empty_df)}** row(s) have an empty ledger name. "
                f"**{cols_txt}** must be filled in for every voucher — a blank "
                "ledger name makes the whole Tally import fail. Fix the rows "
                "below, then re-upload."
            )
            styled_empty = empty_df.style.apply(
                lambda col: [
                    "background-color: #ffd6d6"
                    if (col.name in checked_cols and str(v).strip() == "")
                    else "" for v in col
                ],
                axis=0,
            )
            st.dataframe(styled_empty, use_container_width=True, hide_index=True)
        elif checked_cols:
            st.success(f"✅ No empty {' / '.join(checked_cols)} values.")

    if block_conversion:
        st.info("Conversion is paused until the errors above are resolved.")
        st.stop()

    # ---- Step 2a: Ledger-name check against Tally --------------------------
    convert_bytes = uploaded.getvalue()
    if tool_name == "Purchase / Sales":
        convert_bytes, ledger_blocked = ps_ledger_check_ui(uploaded.getvalue())
        if ledger_blocked:
            st.stop()
    elif tool_name == "Payment / Contra / Receipt":
        convert_bytes, ledger_blocked = pcr_ledger_check_ui(uploaded.getvalue())
        if ledger_blocked:
            st.stop()

    DATE_COLS = {
        "Purchase / Sales": ("Datetime", "Supplier_Invoice"),
        "Payment / Contra / Receipt": ("DATE_TIME", "PartyLedgerName"),
    }

    # ---- Date-format sanity check: warn if Excel auto-converted dates -------
    if tool_name in DATE_COLS:
        date_col, id_col = DATE_COLS[tool_name]
        try:
            conv_df, n_conv = find_excel_converted_dates(
                uploaded.getvalue(), date_col, id_col)
        except Exception:  # noqa: BLE001
            conv_df, n_conv = None, 0
        if conv_df is not None:
            st.warning(
                f"⚠️ **{n_conv}** {date_col} cell(s) were stored by Excel as "
                "**real dates**, not text. If you pasted dates as DD/MM/YYYY, "
                "Excel may have swapped day↔month for days ≤ 12 (e.g. 04/01 "
                "instead of 01/04). To be safe, **re-download the template** "
                "(its date column is now Text) and paste your dates again — "
                "they'll stay exactly as typed. The conversion below reads "
                "these cells as-is."
            )
            with st.expander(f"Show {date_col} cells stored as Excel dates"):
                st.dataframe(conv_df, use_container_width=True, hide_index=True)

    # ---- Step 2b: Voucher date range check -----------------------------
    if date_enabled and tool_name in DATE_COLS:
        date_col, id_col = DATE_COLS[tool_name]
        st.subheader("Step 2b — Voucher date range check")
        if date_from and date_to and date_from <= date_to:
            try:
                oor_df, unparsed_df = find_dates_out_of_range(
                    uploaded.getvalue(), date_from, date_to, date_col, id_col)
            except Exception as exc:  # noqa: BLE001
                oor_df, unparsed_df = None, None
                st.warning(f"Could not validate voucher dates: {exc}")

            span = f"{date_from:%d/%m/%Y} – {date_to:%d/%m/%Y}"
            if oor_df is not None:
                st.error(
                    f"🚫 **{len(oor_df)}** voucher(s) have a date outside the "
                    f"selected range (**{span}**). Fix these rows or widen the "
                    "range, then re-upload."
                )
                st.dataframe(oor_df, use_container_width=True, hide_index=True)
            if unparsed_df is not None:
                st.warning(
                    f"⚠️ **{len(unparsed_df)}** row(s) have a {date_col} that "
                    "couldn't be read — check the date format (DD/MM/YYYY). "
                    "The affected rows are listed below."
                )
                st.dataframe(unparsed_df, use_container_width=True, hide_index=True)
            if oor_df is None and unparsed_df is None:
                st.success(f"✅ All voucher dates fall within {span}.")
            if oor_df is not None:
                st.info("Conversion is paused until the dates above are fixed.")
                st.stop()
        else:
            st.warning("Set a valid From/To date range above to validate "
                       "voucher dates.")

    st.subheader("Step 3 — Convert")
    with st.spinner("Converting to Tally XML…"):
        try:
            # Hand the converter a fresh, seekable copy of the (corrected) bytes.
            result = tool["convert"](io.BytesIO(convert_bytes))

            xml_final = result.get("xml_final") if result else None
            file_name = (result.get("file_name") if result else None) or "tally"

            if not xml_final:
                st.error(
                    "No output was produced. Check that the uploaded file matches "
                    "the template (correct sheet names and column headers)."
                )
            else:
                st.success("Done! Download your files below.")

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.download_button(
                        "⬇️ Tally XML",
                        data=tally_core.result_xml_bytes(result),
                        file_name=f"{file_name}.xml",
                        mime="application/xml",
                    )
                with col2:
                    st.download_button(
                        "⬇️ Excel (.xlsx)",
                        data=tally_core.result_excel_bytes(result),
                        file_name="data.xlsx",
                        mime=XLSX_MIME,
                    )
                with col3:
                    st.download_button(
                        "⬇️ Both (.zip)",
                        data=tally_core.result_zip_bytes(result),
                        file_name="files.zip",
                        mime="application/zip",
                    )

                with st.expander("Preview generated XML"):
                    st.code(xml_final[:5000] + ("\n…" if len(xml_final) > 5000 else ""),
                            language="xml")
        except Exception as exc:  # noqa: BLE001
            st.error(f"Conversion failed: {exc}")
            st.exception(exc)
