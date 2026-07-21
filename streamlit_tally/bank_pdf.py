"""
Bank / loan statement PDF  ->  Payment/Contra/Receipt template, via a VLM.

Every page of the PDF is rendered to an image and sent to a vision model
(OpenAI or Google Gemini), together with the formatting instructions. The model
reads the pages and produces the finished Payment / Contra / Receipt vouchers
directly — inferring the bank and party ledgers from the statement, exactly
like a person filling the template by hand. Because each page is an explicit
image, multi-page statements are read in full.

API keys are read from the environment (a ``.env`` file works): ``OPENAI_API_KEY``
for the gpt-* models, ``GEMINI_API_KEY`` (or ``GOOGLE_API_KEY``) for the
gemini-* models. Pick the model in the UI.
"""
import base64
import io
import json
from typing import List, Literal

import pandas as pd
from pydantic import BaseModel

import tally_core

# Vision-capable models to offer in the UI, across providers.
MODEL_CHOICES = [
    "gemini-3.1-flash-lite", "gemini-2.5-flash", "gemini-2.5-pro",
    "gpt-5-mini", "gpt-5", "gpt-5-nano", "gpt-4o", "gpt-4o-mini",
]
DEFAULT_MODEL = "gemini-2.5-flash"

# Render scale: 72 * SCALE dpi. 2.0 (~144 dpi) keeps statement tables legible
# without blowing up the image-token cost.
RENDER_SCALE = 2.0

# Per-model pricing in USD per 1,000,000 tokens as (input_rate, output_rate).
# These are list prices at the time of writing — edit them if they change. A
# model missing here still shows token counts, just no cost.
PRICING = {
    "gemini-3.1-flash-lite": (0.10, 0.40),  # estimate — verify current rates
    "gemini-2.5-flash": (0.30, 2.50),
    "gemini-2.5-pro": (1.25, 10.00),
    "gpt-5": (1.25, 10.00),
    "gpt-5-mini": (0.25, 2.00),
    "gpt-5-nano": (0.05, 0.40),
    "gpt-4o": (2.50, 10.00),
    "gpt-4o-mini": (0.15, 0.60),
}

# The PCR template columns, in order.
PCR_COLUMNS = [
    "DATE_TIME", "PartyLedgerName", "Dr_LedgerName", "Dr_Amount",
    "Cr_LedgerName", "Cr_Amount", "Narration", "Vch_Type",
    "NARRATION_1", "NARRATION_2",
]

_SYSTEM_PROMPT = (
    "You convert bank and loan account statement pages into Tally accounting "
    "vouchers (Payment, Contra, Receipt) for the account holder's books. "
    "You reply with strict JSON only, matching the provided schema."
)

_INSTRUCTIONS = """\
The images below are the pages of a single bank / loan account statement, in \
order. Read EVERY page and convert EVERY transaction row in the ledger table \
into a Tally voucher. Produce one voucher per transaction, in statement order. \
The transactions span multiple pages — do not stop after the first page.

Classification depends ONLY on which amount column the value sits in — never on \
the wording of the particulars:
- An amount in the DEBIT (DR) column is ALWAYS a **Payment**, even when the \
description contains the word "Receipt" (e.g. "Amount Paid Vide Receipt No." is \
a Payment because its amount is in the DR column).
- An amount in the CREDIT (CR) column is ALWAYS a **Receipt**, even when the \
description contains the word "Paid", "Payment" or "Disbursement".
- Exception: a transaction settled in **cash** (cash deposit, cash withdrawal, \
cash payment) becomes a **Contra**.
Do NOT infer the voucher type from words like "paid", "payment", "receipt" or \
"disbursement" in the description — use the DR / CR column only.

Ledger names — infer them from the statement itself, do not invent generic ones:
- BANK ledger: the bank / finance-company account the statement is for (use the \
bank name, or for a loan statement the lender / finance company name).
- PARTY ledger: the counterparty for the transaction. On a single-counterparty \
statement (e.g. a loan account) use that lender's name for every voucher.
- For cash transactions use a ledger named "Cash".
- If the statement shows an account or loan number, include it in the narration.

Fill ALL of these fields for every voucher:
- DATE_TIME: the transaction date as dd/mm/yyyy (day first).
- Vch_Type: exactly one of "Payment", "Receipt", or "Contra".
- Dr_Amount and Cr_Amount: both equal the transaction amount (one positive \
number, no commas or currency symbols).
- Payment: Dr_LedgerName = party ledger, Cr_LedgerName = bank ledger.
- Receipt: Dr_LedgerName = bank ledger, Cr_LedgerName = party ledger.
- Contra: put "Cash" on the correct side and the bank/party ledger on the other.
- PartyLedgerName: the counterparty ledger name.
- Narration: a clear description of the voucher, built from the statement \
particulars.
- NARRATION_1: narration for the debit line. NARRATION_2: narration for the \
credit line.

Do not skip any transaction row on any page. Do not include opening/closing \
balance summary lines. Return only JSON.
"""

# --- OpenAI: JSON-schema for strict structured output --------------------
_OPENAI_SCHEMA = {
    "type": "object",
    "properties": {
        "vouchers": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "DATE_TIME": {"type": "string"},
                    "PartyLedgerName": {"type": "string"},
                    "Dr_LedgerName": {"type": "string"},
                    "Dr_Amount": {"type": "number"},
                    "Cr_LedgerName": {"type": "string"},
                    "Cr_Amount": {"type": "number"},
                    "Narration": {"type": "string"},
                    "Vch_Type": {"type": "string",
                                 "enum": ["Payment", "Contra", "Receipt"]},
                    "NARRATION_1": {"type": "string"},
                    "NARRATION_2": {"type": "string"},
                },
                "required": PCR_COLUMNS,
                "additionalProperties": False,
            },
        },
    },
    "required": ["vouchers"],
    "additionalProperties": False,
}


# --- Gemini: pydantic response schema ------------------------------------
class _Voucher(BaseModel):
    DATE_TIME: str
    PartyLedgerName: str
    Dr_LedgerName: str
    Dr_Amount: float
    Cr_LedgerName: str
    Cr_Amount: float
    Narration: str
    Vch_Type: Literal["Payment", "Contra", "Receipt"]
    NARRATION_1: str
    NARRATION_2: str


class _Vouchers(BaseModel):
    vouchers: List[_Voucher]


def provider_for(model):
    """Return the provider ('gemini' or 'openai') for a model name."""
    return "gemini" if str(model).lower().startswith("gemini") else "openai"


def _pdf_to_page_pngs(pdf_bytes, scale=RENDER_SCALE):
    """Render every PDF page to PNG bytes (one per page, in order)."""
    import pypdfium2 as pdfium  # lazy: optional dependency

    pages = []
    pdf = pdfium.PdfDocument(pdf_bytes)
    try:
        for i in range(len(pdf)):
            pil = pdf[i].render(scale=scale).to_pil()
            buf = io.BytesIO()
            pil.save(buf, format="PNG")
            pages.append(buf.getvalue())
    finally:
        pdf.close()
    return pages


def _vouchers_openai(page_pngs, model, api_key):
    """Extract voucher dicts from page images using an OpenAI vision model."""
    from openai import OpenAI  # lazy: optional dependency

    content = [{"type": "text", "text": _INSTRUCTIONS}]
    for png in page_pngs:
        b64 = base64.standard_b64encode(png).decode("utf-8")
        content.append({"type": "image_url", "image_url": {
            "url": f"data:image/png;base64,{b64}", "detail": "high"}})

    client = OpenAI(api_key=api_key) if api_key else OpenAI()
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user", "content": content},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {"name": "pcr_vouchers", "strict": True,
                            "schema": _OPENAI_SCHEMA},
        },
    )
    data = json.loads(response.choices[0].message.content or "{}")
    u = response.usage
    usage = {
        "input_tokens": getattr(u, "prompt_tokens", 0) or 0,
        "output_tokens": getattr(u, "completion_tokens", 0) or 0,
    }
    return data.get("vouchers", []), usage


def _vouchers_gemini(page_pngs, model, api_key):
    """Extract voucher dicts from page images using a Google Gemini model."""
    from google import genai  # lazy: optional dependency
    from google.genai import types

    client = genai.Client(api_key=api_key) if api_key else genai.Client()
    parts = [types.Part.from_bytes(data=png, mime_type="image/png")
             for png in page_pngs]
    response = client.models.generate_content(
        model=model,
        contents=[_INSTRUCTIONS, *parts],
        config=types.GenerateContentConfig(
            system_instruction=_SYSTEM_PROMPT,
            response_mime_type="application/json",
            response_schema=_Vouchers,
        ),
    )
    data = json.loads(response.text or "{}")
    um = getattr(response, "usage_metadata", None)
    inp = getattr(um, "prompt_token_count", 0) or 0 if um else 0
    tot = getattr(um, "total_token_count", 0) or 0 if um else 0
    # total includes any thinking tokens; output = everything that isn't input.
    out = (tot - inp) if tot else (getattr(um, "candidates_token_count", 0) or 0)
    usage = {"input_tokens": inp, "output_tokens": max(out, 0)}
    return data.get("vouchers", []), usage


def _row(v):
    """Normalise one raw voucher dict to the PCR columns."""
    return {
        "DATE_TIME": _norm_date(v.get("DATE_TIME", "")),
        "PartyLedgerName": v.get("PartyLedgerName", ""),
        "Dr_LedgerName": v.get("Dr_LedgerName", ""),
        "Dr_Amount": _num(v.get("Dr_Amount", 0)),
        "Cr_LedgerName": v.get("Cr_LedgerName", ""),
        "Cr_Amount": _num(v.get("Cr_Amount", 0)),
        "Narration": v.get("Narration", ""),
        "Vch_Type": v.get("Vch_Type", ""),
        "NARRATION_1": v.get("NARRATION_1", ""),
        "NARRATION_2": v.get("NARRATION_2", ""),
    }


def _norm_date(value):
    """Normalise any date the model returns to dd/mm/yyyy (day-first)."""
    ts = tally_core.suggest_date(value)
    return ts.strftime("%d/%m/%Y") if ts is not None else str(value)


def _num(value):
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def pcr_from_pdf(pdf_bytes, model=DEFAULT_MODEL, api_key=None, scale=RENDER_SCALE):
    """Convert a statement PDF into a PCR-format DataFrame using a vision model.

    Every page is rendered to an image and sent to the model (OpenAI or Gemini,
    chosen by the model name). Returns ``(pcr_df, usage)`` where usage carries
    token counts and an estimated cost.
    """
    page_pngs = _pdf_to_page_pngs(pdf_bytes, scale=scale)
    if provider_for(model) == "gemini":
        vouchers, usage = _vouchers_gemini(page_pngs, model, api_key)
    else:
        vouchers, usage = _vouchers_openai(page_pngs, model, api_key)

    df = pd.DataFrame([_row(v) for v in vouchers], columns=PCR_COLUMNS)
    usage.update(_usage_extras(model, usage["input_tokens"], usage["output_tokens"]))
    return df, usage


def _usage_extras(model, input_tokens, output_tokens):
    """Add model, total tokens, per-1M rates and estimated USD cost."""
    rates = PRICING.get(model)
    extras = {
        "model": model,
        "pages": None,
        "total_tokens": input_tokens + output_tokens,
        "input_rate": rates[0] if rates else None,
        "output_rate": rates[1] if rates else None,
        "input_cost": None, "output_cost": None, "cost_usd": None,
    }
    if rates:
        extras["input_cost"] = input_tokens / 1_000_000 * rates[0]
        extras["output_cost"] = output_tokens / 1_000_000 * rates[1]
        extras["cost_usd"] = extras["input_cost"] + extras["output_cost"]
    return extras


def pcr_to_xlsx(pcr_df):
    """Write a PCR-format DataFrame into the blank PCR template (Text dates)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(tally_core.template_pay_con_rec()))
    ws = wb["TEMPLATE"]
    headers = [c.value for c in ws[1]]
    for r, row in enumerate(pcr_df.to_dict("records"), start=2):
        for header, value in row.items():
            if header in headers:
                ws.cell(row=r, column=headers.index(header) + 1, value=value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
